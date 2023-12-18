import os
import re
import random
import pathlib
import zipfile
import datetime
import copy
import logging
import pandas as pd
import numpy as np
import scipy as sp
from scipy import interpolate
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from pycel import ExcelCompiler
from openpyxl import Workbook, load_workbook
from anytree import Node, find, findall
from copy import copy, deepcopy
from faker import Faker
from .Helper import Helper
from .OutputAnalyzer import OutputAnalyzer

class SheetOutputGenerator:
    """
    The SheetOutputGenerator class is responsible for generating the final output Excel file by replacing variables in the output sheets with their corresponding values obtained from the interpreter.

    Args:
        interpreter (Interpreter): The interpreter object used to obtain the values for the variables in the output sheets.
        output_path (str): The path to the output files.

    Attributes:
        output_path (str): The path to the output files.
        interpreter (Interpreter): The interpreter object used to obtain the values for the variables in the output sheets.
        all_sheets (dict): A dictionary containing the analyzed output sheets for each output file.
    """
    def __init__(self, interpreter, output_path):
        """
        Initializes the SheetOutputGenerator object with the specified interpreter and output path.

        Args:
            interpreter (Interpreter): The interpreter object used to obtain the values for the variables in the output sheets.
            output_path (str): The path to the output files.
        """
        self.output_path = output_path
        self.interpreter = interpreter
        self.all_sheets = None

    def analyzeAllOutputSheet(self):
        """
        Analyzes all output sheets by creating an OutputAnalyzer object for each sheet in each output file.
        Sets self.all_sheets to a dictionary containing the analyzed output sheets for each output file.
        """
        all_files = next(os.walk(self.output_path), (None, None, []))[2]
        all_wks = {file: load_workbook(self.output_path + file) for file in all_files if file.endswith('.xlsx')}
        
        self.all_sheets = {}
        for file, wb in all_wks.items():
            sheetsDic = {
                sheet_name: OutputAnalyzer(wb, sheet_name, self.output_path + file, self.interpreter)
                for sheet_name in all_wks[file].sheetnames
            }
            self.all_sheets[file] = sheetsDic

    def save_data_to_excel(self, data, file_name):
        wb = Workbook()
        # Création des feuilles et insertion des données
        for category, category_data in data.items():

            ws = wb.create_sheet(title=category)

            # Création des colonnes pour chaque type de données
            columns = {}
            for item in category_data['curves'] + category_data['constants']:

                item_name = item['curve_name'] if 'curve_name' in item.keys() else item['constant_name']
                if 'values' in item.keys():
                    columns[item_name] = item['values']
                else:
                    columns[item_name] = [item['value']]
                                
            # Ajout des en-têtes de colonnes et des données
            ws.append(list(columns.keys()))
            max_length = max(len(col) for col in columns.values())

            for i in range(max_length):
                row = [columns[name][i] if i < len(columns[name]) else '' for name in columns]
                ws.append(row)

        # Suppression de la feuille par défaut
        del wb['Sheet']
        wb.save(file_name)

    def generate(self, folder, zip_fn):
        """
        Generates the final output Excel file by replacing variables in the output sheets with their corresponding values obtained from the interpreter.

        Args:
            folder (str): The folder to save the generated output files.
            zip_fn (str): The filename of the generated zip file.

        Returns:
            str: The path to the generated zip file.
        """
        from pathlib import Path
        os.makedirs(folder, exist_ok=True)

        data_matrix = self.interpreter.tree.generate_tree_dict(self.interpreter.tree.root)
        self.save_data_to_excel(data_matrix, f"{folder}/data.xlsx")

        count = 0
        for fPath, sheets in self.all_sheets.items():
            count +=1
            fn = Path(fPath).stem
            for sheet_name, analyzer in sheets.items():
                analyzer.findAndReplaceAnnotateValues()
                analyzer.save(f"{folder}/{fn}.xlsx")

        return Helper.folder_zip(folder, zip_fn)
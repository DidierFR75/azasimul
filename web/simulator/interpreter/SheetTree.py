import os
import pandas as pd
import numpy as np
import scipy as sp
from scipy import interpolate
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from pycel import ExcelCompiler
from openpyxl import Workbook, load_workbook
from anytree import Node, find, findall
from copy import copy, deepcopy
from faker import Faker
from .InputAnalyzer import InputAnalyzer
from .Helper import Helper

import logging
import os
try:
    logging.config.dictConfig({
    'version': 1,
    'disable_existing_loggers': False,
    'formatters': {
        'console': {
            'format': '%(name)-12s %(levelname)-8s %(message)s'
        },
        'file': {
            'format': '%(asctime)s %(name)-12s %(levelname)-8s %(message)s'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'formatter': 'console'
        },
        'file': {
            'level': 'DEBUG',
            'class': 'logging.FileHandler',
            'formatter': 'file',
            'filename': './logs/log.log'
        }
    },
    'loggers': {
        '': {
            'level': 'DEBUG',
            'handlers': ['console', 'file']
        }
    }
})
    os.makedirs('./logs', exist_ok=True)
    logger = logging.getLogger(__name__)
except Exception as e:
    print(e)

class SheetTree:
    """
    The `SheetTree` class is responsible for creating a tree structure to organize and analyze sheets in Excel workbooks.

    Attributes:
        path (str): The path to the folder containing the Excel workbooks.
        root (Node): The root node of the formula tree.
        all_sheet (dict): A dictionary with the file names as keys and a list of sheet analyzers as values.
        operation_sheets (list): A list of sheet analyzers for the operation sheets.
    """

    SUMMARY_SHEET = "UseCase.xlsx"

    def __init__(self, path) -> None:
        """
        Initializes a new instance of the SheetTree class.
        
        Args:
            path (str): The path to the folder containing the Excel workbooks.
        """
        self.path = path
        
        self.root = Node("root")
        self.root.name = "Summary"
        self.root.categories = {}

        self.all_sheet = None
        self.operation_sheets = []

        self.num_points = 0
   
    def readAllSheetsFromFolder(self, folder):
        """
        Reads all the sheets from a folder and returns a dictionary with the file names as keys and a list of sheet analyzers as values.
        
        Args:
            folder (str): The path to the folder containing the Excel workbooks.
        
        Returns:
            dict: A dictionary with the file names as keys and a list of sheet analyzers as values.
        """
        result = {}
        main_summary = []

        # Load all workbooks
        all_files = next(os.walk(folder), (None, None, []))[2]
        all_files = [ fn for fn in all_files if not Helper.rejectXlsFile(fn) ]

        if self.SUMMARY_SHEET in all_files:
            all_files.remove(self.SUMMARY_SHEET)
            all_files.insert(0, self.SUMMARY_SHEET)

        all_wks = { file: load_workbook(folder +'/'+ file) for file in all_files }
        
        # Create dict with file: {sheetname: analyzer}
        test = []
        for file, wb in all_wks.items():
            result[file] = []
            for sheet_name in wb.sheetnames:
                if sheet_name.startswith(Helper.DELIMITER_SHEET_UNFOLLOW):
                    logger.info(f"Sheet('{sheet_name}') : SKIPPED")
                    continue
                analyzer = InputAnalyzer(wb[sheet_name], sheet_name, folder + '/' + file, main_summary)
                if analyzer.loadSheet():
                    result[file].append((sheet_name, analyzer,))
                    if file == self.SUMMARY_SHEET:
                        main_summary = analyzer.summary
                    if analyzer.num_points != 0:
                        self.num_points = analyzer.num_points
    
        return result
    
    def mapSheetsToFormulaTree(self, path=None):
        """
        Maps the sheets to the formula tree structure.
        
        Args:
            path (str, optional): The path to the folder containing the Excel workbooks. If not provided, uses the default path.
        """
        if not path:
            path = self.path
        nodes = []
        self.all_sheet = self.readAllSheetsFromFolder(path)
        # Create all nodes
        for _file, wbSheets in self.all_sheet.items():
            for sheet_name, analyzer in wbSheets:
                
                if analyzer.isOperationSheet():
                    self.operation_sheets.append(analyzer)
                    continue
            
                if analyzer.isConstantSheet():
                    Node(sheet_name, analyzer=analyzer, parent=self.root)
                    continue

                if analyzer.isSummarySheet():
                    if not hasattr(self.root, "analyzer"):
                        self.root.analyzer = analyzer                                    
                    continue

                if analyzer.metadatas == {}:
                    continue
                
                # last Case is a "Curves" sheet
                # Get parent name if exists
                parentName = analyzer.metadatas[analyzer.PRODUCT_PARENT] if (analyzer.PRODUCT_PARENT in analyzer.metadatas) else None
                    
                productType = analyzer.metadatas[analyzer.PRODUCT_NAME]
                node = Node(productType, analyzer=analyzer)
                
                # Get and add category to self.root.categories if exist
                category = analyzer.getCategory()
                if category is not None:
                    self.root.categories[category.lower()] = category.upper()+":"
                    node.category = category.lower()
                                            
                nodes.append( (parentName, productType, node) )
        
        # Add parent for all nodes
        for element in nodes:
            if element[0] is None: # if no parentName
                element[2].parent = self.root
            else:
                i = [i for i, v in enumerate(nodes) if v[1] == element[0] and element[2].category == v[2].category]
                if i != []:
                    element[2].parent = nodes[i[0]][2]

    def generate_tree_dict(self, node):
        data_by_category = {}

        def add_data(child_node):
            category = child_node.category if hasattr(child_node, 'category') and child_node.category is not None else 'root'
            if not category in data_by_category:
                data_by_category[category] = {"curves": [], "constants": [], "operations": []}

            if hasattr(child_node, 'analyzer'):
                # Traiter les courbes
                curves = getattr(child_node.analyzer, 'curves', [])
                for curve in curves:
                    #if curve.get('interpolation') == 'CONST':
                    #    curve['values'] = [curve['values'][0]]
                    curve_data = {k: v for k, v in curve.items() if k not in ['column', 'interpolation', "specificiation_name"]}
                    data_by_category[category]["curves"].append(curve_data)

                # Traiter les constantes
                constants = getattr(child_node.analyzer, 'constants', {})
                for const_key, const_values in constants.items():
                    for const in const_values:
                        data_by_category[category]["constants"].append(const)

                # Traiter les opérations
                operations = getattr(child_node.analyzer, 'operations', {})
                for op_key, op_values in operations.items():
                    for op in op_values:
                        data_by_category[category]["operations"].append(op)

            for child in child_node.children:
                add_data(child)

        add_data(node)

        return data_by_category




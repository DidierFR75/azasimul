import pandas as pd
import numpy as np
import scipy as sp
from scipy import interpolate
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from pycel import ExcelCompiler
from openpyxl import Workbook, load_workbook
from anytree import Node, find, findall
from copy import copy, deepcopy
from faker import Faker
from .InputAnalyzer import InputAnalyzer

class FileChecker:
    """
    This class is responsible for checking the format of a file. 
    It loads the file using the `load_workbook` function from the `openpyxl` library and iterates over each sheet in the workbook. 
    For each sheet, it creates an instance of the `InputAnalyzer` class to analyze the sheet. 
    If the sheet is a summary sheet, it stores the summary data. 
    If the sheet is a constant or operation sheet, it adds the sheet name to the list of non-accepted files. 
    Finally, it saves the modified workbook.

    Fields:
    - path: The path of the file to be checked.
    - summary: The summary data extracted from the summary sheet.
    - non_accepted: A list of sheet names that are not accepted (constant or operation sheets).
    - wb: The `Workbook` object representing the loaded workbook.
    """
    def __init__(self, path) -> None:
        """
        Initializes the `FileChecker` object with the path of the file to be checked.

        Args:
        - path: The path of the file to be checked.
        """
        self.path = path
        self.summary = None
        self.non_accepted = []
        self.wb = None

    def checkForSpecFormat(self):
        """
        Checks the format of the file. Loads the file using `load_workbook` function, iterates over each sheet, analyzes each sheet using `InputAnalyzer` class, stores the summary data if the sheet is a summary sheet, adds the sheet name to the list of non-accepted files if the sheet is a constant or operation sheet, and saves the modified workbook.

        :return: None
        """
        self.wb = load_workbook(self.path)
        for sheet_name in self.wb.sheetnames:
            analyzer = InputAnalyzer(self.wb[sheet_name], sheet_name, self.path)
            if analyzer.loadSheet():
                if analyzer.isSummarySheet():
                    self.summary = analyzer.summary
                    self.wb.remove(self.wb[sheet_name])
            
                if analyzer.isConstantSheet() or analyzer.isOperationSheet():
                    self.non_accepted.append(sheet_name)
                    self.wb.remove(self.wb[sheet_name])         
        self.wb.save(self.path)
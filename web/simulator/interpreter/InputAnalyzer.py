import random
import datetime
import pandas as pd
import numpy as np
import scipy as sp
from scipy import interpolate
from scipy.stats import linregress
from scipy.optimize import curve_fit
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from pycel import ExcelCompiler
from openpyxl import Workbook, load_workbook
from anytree import Node, find, findall
from copy import copy, deepcopy
from faker import Faker

import logging
import os
try:
    logging.config.dictConfig({
    'version': 1,
    'disable_existing_loggers': False,
    'formatters': {
        'console': {
            'format': '%(name)-12s %(levelname)-8s %(message)s'
        },
        'file': {
            'format': '%(asctime)s %(name)-12s %(levelname)-8s %(message)s'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'formatter': 'console'
        },
        'file': {
            'level': 'DEBUG',
            'class': 'logging.FileHandler',
            'formatter': 'file',
            'filename': './logs/log.log'
        }
    },
    'loggers': {
        '': {
            'level': 'DEBUG',
            'handlers': ['console', 'file']
        }
    }
})
    os.makedirs('./logs', exist_ok=True)
    logger = logging.getLogger(__name__)
except Exception as e:
    print(e)
    
class InputAnalyzer:
    """
    The `InputAnalyzer` class is responsible for analyzing and processing data from an Excel worksheet. 
    It can identify different types of sheets (operation, constant, summary, and curves) and extract relevant information from each type. 
    It can also perform interpolation on curves and store the results.

    Main functionalities:
    - Identify and load different types of sheets (operation, constant, summary, and curves)
    - Extract metadata, points, curves, constants, summary, and operations from the sheet
    - Perform interpolation on curves
    - Add new curves, constants, summary, and operations to the analyzer

    Methods:
    - loadSheet(): Loads and analyzes the sheet, extracting relevant information based on the sheet type
    - addCurve(specification_name, value, unit, interpolation): Adds a new curve to the analyzer
    - addSummary(name, value, unit): Adds a new summary to the analyzer
    - getCurveByName(name): Retrieves a curve by its name
    - getConstantByName(name): Retrieves a constant by its name
    - getSummaryByName(name): Retrieves a summary by its name
    - getOperationByName(name): Retrieves an operation by its name
    - getCategory(): Retrieves the category metadata of the sheet

    Fields:
    - BASE_ELEMENTS_ROW: The row number where the base elements are located
    - UNIT_ROW: The row number where the units are located
    - CURVE: The row number where the curve interpolations are located
    - METADATA_COL: The column letter where the metadata is located
    - POINT_N_COL: The column letter where the point numbers are located
    - DATE_COL: The column letter where the dates are located
    - PRODUCT_NAME: The metadata name for the product type
    - PRODUCT_PARENT: The metadata name for the product subtype
    - CATEGORY: The metadata name for the category
    - DELIMITER_SHEET_UNFOLLOW: The delimiter used to skip sheets
    - CONSTANT_SHEETNAME: The name of the constant sheet
    - SUMMARY_SHEETNAME: The name of the summary sheet
    - OPERATION_SHEETNAME: The name of the operation sheet
    - evaluator: An instance of the ExcelCompiler class for evaluating Excel formulas
    - sheet_name: The name of the sheet being analyzed
    - ws: The worksheet object being analyzed
    - curves: A list of dictionaries representing the curves in the sheet
    - points: A list of dictionaries representing the points in the sheet
    - metadatas: A dictionary representing the metadata in the sheet
    - operations: A dictionary representing the operations in the sheet
    - constants: A dictionary representing the constants in the sheet
    - summary: A list of dictionaries representing the summary in the sheet
    - fake: An instance of the Faker class for generating fake data
    """
    BASE_ELEMENTS_ROW = 17 # Location of base's elements
    UNIT_ROW = 18 # Location of units
    CURVE_ROW = 19 # Location of Curve's interpolations
    METADATA_COL = "A"
    POINT_N_COL = "B"
    DATE_COL = "D"
    PRODUCT_NAME = "Product-Type"
    PRODUCT_PARENT = "SubType"
    CATEGORY = "Category"
    DELIMITER_SHEET_UNFOLLOW = "_"

    CONSTANT_SHEETNAME = "Constants"
    SUMMARY_SHEETNAME = "Summary" 
    OPERATION_SHEETNAME = "operation"

    SIMULATION_FREQUENCY_NAME = 'Simulation Frequency'
    SIMULATION_STARTDATE_NAME = "Start"
    SIMULATION_END_NAME = "End"
    FREQ_MULTIPLIER = {
        'year': 1,
        'semester': 2,
        'quarter': 4,
        'month': 12,
        'week': 52,
        'day': 365
    }
    PD_FREQ_MULTIPLIER = {
        'year': 'A',
        'semester': '6M',
        'quarter': '3M',
        'month': 'M',
        'week': 'W',
        'day': 'D'
    }

    def __init__(self, ws, sheet_name, path, summary=[]) -> None:
        self.evaluator = ExcelCompiler(filename=path)
        self.sheet_name = sheet_name
        self.ws = ws

        self.curves = []
        self.points = []
        self.num_points = 0
        self.duration_in_years = 0
        self.metadatas = {}
        self.operations = {}
        self.constants = {}
        self.summary = summary
        self.fake = Faker()
    
    ### Util's functions

    def log_interp1d(self, xx, yy, kind='linear', small_value=1e-10):
        """
        Return the log interpolation on 1 dimension

        Perform logarithmic interpolation on one-dimensional data.

        Parameters:
        xx (array): The x values of the data points.
        yy (array): The y values of the data points.
        kind (string, optional): The type of interpolation to perform. Defaults to 'linear'.
        small_value (float, optional): A small value used to adjust the y values to avoid taking the logarithm of zero. Defaults to 1e-10.

        Returns:
        log_interp (lambda function): A lambda function that can be used to interpolate new y values based on the given x and y values.
        """
        yy_adjusted = np.maximum(yy, small_value)
        logx = np.log10(xx)
        logy = np.log10(yy_adjusted)
        lin_interp = interpolate.interp1d(logx, logy, kind=kind, fill_value="extrapolate")
        log_interp = lambda zz: np.power(10.0, lin_interp(np.log10(zz)))
        return log_interp
    
    def determine_interpolation_type(self, values):
        values = [float(v) for v in values]
        x = np.arange(len(values))

        # Checks if all values are constant
        if all(value == values[0] for value in values):
            return "CONST"

        # Calculating R^2 to evaluate fit
        def calculate_r_squared(y_true, y_fit):
            ss_res = np.sum((y_true - y_fit) ** 2)
            ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
            return 1 - (ss_res / ss_tot)

        # Linear interpolation
        slope, intercept, r_value, _, _ = linregress(x, values)
        if r_value**2 > 0.95:
            return "LINEAR"

        # Other types of interpolation
        interpolations = {
            "LOG": lambda x, a, b: a * np.log(x) + b,
            "EXP": lambda x, a, b: a * np.exp(b * x),
            "SIN": lambda x, a, b, c: a * np.sin(b * x) + c,
            # Add here other types of interpolation with their corresponding functions
        }

        for name, func in interpolations.items():
            try:
                popt, _ = curve_fit(func, x, values, maxfev=10000)
                if calculate_r_squared(values, func(x, *popt)) > 0.95:
                    return name
            except:
                pass

        return "Unknown"


    def evaluate(self, cell):
        """
        Return the evaluation's value of a given cell (rounded by 3)

        Args:
            cell (str): The cell reference in the Excel sheet (e.g., "A1").

        Returns:
            float: The evaluation value of the given cell, rounded to three decimal places.
        """
        eval = self.evaluator.evaluate(self.sheet_name+"!"+cell.coordinate)
        if isinstance(eval, float):
            eval = round(eval, 3)
        return eval
    
    def _clean_string(self, text):
        return str(text).replace("\n", "").replace("\t", "").lstrip()

    ### 4 Sheet types

    def isOperationSheet(self):
        if self.OPERATION_SHEETNAME.lower() in self.sheet_name.lower():
            return True
        return False

    def isConstantSheet(self):
        if self.CONSTANT_SHEETNAME.lower() in self.sheet_name.lower():
            return True
        return False
    
    def isSummarySheet(self):
        if self.SUMMARY_SHEETNAME.lower() in self.sheet_name.lower():
            return True
        return False

    def isCurvesSheet(self):
        if self.curves != [] and self.points != []:
            return True
        return False

    # Generator's functions

    def _populate_by_frequency(self, points, X=1):
        """
        Populates the given list of points with interpolated dates based on a reference date.

        Parameters:
        - points (list): A list of dictionaries where each dictionary has a "date" key.
                        The list should have at least one point with a non-null date.
        - granularity (str): The unit for interpolation. Supported values are now "year", "semester",
                        "quarter", "month", "week", "day", "hour", and "minute".
        - X (int): The multiplier for the granularity.

        The function will find the first non-null date in the points list and use it as a reference.
        It will then populate missing dates in the list based on the given granularity and multiplier.

        Returns:
        - None: The function modifies the input list in-place.
        """

        # Extended granularity map to include semester, quarter, and week
        granularity_map = {
            "year": lambda base_date, diff: base_date + relativedelta(years=random.randint(1, diff)),
            "semester": lambda base_date, diff: base_date + relativedelta(months=6*random.randint(1, diff)),
            "quarter": lambda base_date, diff: base_date + relativedelta(months=3*random.randint(1, diff)),
            "month": lambda base_date, diff: base_date + relativedelta(months=random.randint(1, diff)),
            "week": lambda base_date, diff: base_date + timedelta(weeks=random.randint(1, diff)),
            "day": lambda base_date, diff: base_date + timedelta(days=random.randint(1, diff)),
            "hour": lambda base_date, diff: base_date + timedelta(hours=random.randint(1, diff)),
            "minute": lambda base_date, diff: base_date + timedelta(minutes=random.randint(1, diff))
        }

        # Create the reference date as the first one found
        ref_date = next(({"index": index, "date": point["date"]} for index, point in enumerate(points) if point["date"]), None)

        if ref_date:
            for index, point in enumerate(points):
                if point["date"] is None:
                    point["date"] = granularity_map[self.getSimulationFrequency()](ref_date["date"], X)
        
        return points

    def _generate_points_with_dates(self):
        """
        Return a dictionary of points with their associated dates.

        This method collects the row number, point number, and date for each point in the worksheet.
        It filters out any points that do not have a date.
        If there are at least two points with dates, it performs linear interpolation to calculate the dates for the remaining points.
        If there are fewer than two points with dates, it populates the missing dates based on a specified granularity.

        Returns:
        points_with_dates (dict): A dictionary containing the row number, point number, and date for each point in the worksheet.
        """

        points = [
            {"row": point.row, "point_n": self.evaluate(point), "date": self.ws[self.DATE_COL+str(point.row)].value } 
            for point in self.ws[self.POINT_N_COL] 
            if point.value is not None and (isinstance(point.value, (int, float)) or (isinstance(point.value, str) and point.value.startswith("=")))
        ]
        valid_points = [ pt for pt in points if pt["date"] is not None ]
        
        # Check if exist at least 2 dates in points
        if len(valid_points) >= 2:
            x = [points.index(pt) for pt in valid_points]
            y = [pt["date"].timestamp() for pt in valid_points]

            interp1d = interpolate.interp1d(x, y, fill_value="extrapolate")
            interpolated_timestamps = interp1d(range(self.num_points))

            limited_points = points[:self.duration_in_years]

            for pt, ts in zip(limited_points, interpolated_timestamps):
                if pt["date"] is None:
                    pt["date"] = datetime.fromtimestamp(ts)
            
            return limited_points
        else:
            return self._populate_by_frequency(points, self.getSimulationFrequency())

    def _generateCurves(self):
        """
        Return a sorted (by ASC) dict (column, curve, unit, interpolation) at BASE_ELEMENTS_ROW for a given sheetname
        Ignore the firsts 2 elements because they always not belong to Curves
        """
        return sorted([{
            "column": be.column, 
            "curve_name": self._clean_string(be.value),
            "values": None,
            "unit": self.ws.cell(row=self.UNIT_ROW, column=be.column).value, 
            "interpolation": self.ws.cell(row=self.CURVE_ROW, column=be.column).value
        } for be in self.ws[self.BASE_ELEMENTS_ROW] if be.value is not None][2:], key=lambda x: x["column"])

    def _generateMetaData(self):
        """
        Return array of tuples (row_id, metadata_name, metadata_value) for a given sheetname
        """
        return {self._clean_string(cell.value): self.ws['B'][cell.row-1].value 
            for cell in self.ws[self.METADATA_COL] 
            if (cell.value is not None and self.ws['B'][cell.row-1].value is not None and cell.row < self.BASE_ELEMENTS_ROW)}

    def _generateConstants(self):
        """
        Return all global constant in the sheet
        """
        compositions = [(cmp.row, cmp.value) for cmp in self.ws["A"] if cmp.value is not None]

        result = {}
        for index, cmp in enumerate(compositions):
            tmp = []            
            if(index != len(compositions)-1):
                last_row = compositions[index+1][0]-1
            else:
                last_row = len(self.ws["B"])

            for x in range(cmp[0]+1, last_row):
                if self.ws["B"+str(x)].value is not None:
                    tmp.append({
                        "constant_name": self._clean_string(self.ws["B"+str(x)].value),
                        "value": self.evaluate(self.ws["C"+str(x)]),
                        "unit": self.evaluate(self.ws["D"+str(x)])
                    })
            
            result[cmp[1]] = tmp

        return result

    def _generateSummary(self):
        for composition in self.ws["A"]:
            if composition.value is not None and self.ws.cell(row=composition.row, column=composition.column+1).value is not None:
                self.addSummary(composition.value, self.ws.cell(row=composition.row, column=composition.column+1).value, self.ws.cell(row=composition.row, column=composition.column+2).value)

    def _generateOperations(self):
        """
        Extracts and generates a dictionary of operations from the Excel worksheet.

        Returns:
            dict: A dictionary where the keys are the operation names and the values are lists of dictionaries containing the operation details.
        """
        items = [(it.row, it.value) for it in self.ws["A"] if it.value is not None]

        result = {}
        for index, fcn in enumerate(items):
            tmp = []            
            if(index != len(items)-1):
                last_row = items[index+1][0]-1
            else:
                last_row = len(self.ws["B"])

            for x in range(fcn[0]+1, last_row):
                if self.ws["B"+str(x)].value is not None:
                    tmp.append({
                        "operation_name": self._clean_string(self.ws["B"+str(x)].value),
                        "operation": self.evaluate(self.ws["C"+str(x)]),
                        "unit": self._clean_string(self.ws["D"+str(x)].value) 
                    })
        
            result[fcn[1]] = tmp

        return result

    def _determine_current_frequency(self, points):
        """
        Determines the closest frequency based on the average time difference between consecutive points in a time series.

        Args:
            points (list): A list of dictionaries representing the points in a time series, where each dictionary has a "date" key.

        Returns:
            str: The determined frequency based on the average time difference between consecutive points.
        """
        if len(points) < 2:
            return 'year'  # Default if not enough points to determine

        # Calculate time differences between consecutive points
        date_diffs = [points[i+1]['date'] - points[i]['date'] for i in range(len(points)-1)]
        avg_diff = sum(date_diffs, timedelta(0)) / len(date_diffs)

        # Dictionary to associate the average time difference with a frequency
        diff_to_freq = {
            'year': timedelta(days=365),
            'semester': timedelta(days=182.5),
            'quarter': timedelta(days=91.25),
            'month': timedelta(days=30.44),
            'week': timedelta(days=7),
            'day': timedelta(days=1)
        }

        # Find the frequency closest to the mean difference
        closest_frequency = min(diff_to_freq, key=lambda k: abs(diff_to_freq[k] - avg_diff))

        return closest_frequency

    def _adjust_points_frequency(self, points, values,   frequency):
        # Créer une plage de dates de start_year à end_year à la fréquence spécifiée
        freq_rule = self.PD_FREQ_MULTIPLIER[frequency]
        date_range = pd.date_range(
            start=self.getSummaryByName(self.SIMULATION_STARTDATE_NAME)["summary_value"],
            end=self.getSummaryByName(self.SIMULATION_END_NAME)["summary_value"],
            freq=freq_rule
        )
        
        # Convertir les points et les valeurs en DataFrame Pandas
        df = pd.DataFrame(points)
        df_values = pd.DataFrame(values)

        # Fusionner df et df_values en incluant la colonne 'date'
        df['date'] = pd.to_datetime(df['date'])
        df_values['date'] = df['date']
        df_full = pd.merge(df, df_values, on=['row', 'date'], how='left')

        # Interpoler ou étendre les points en utilisant la plage de dates
        df_full.set_index('date', inplace=True)
        df_full = df_full.reindex(date_range, method='pad')

        # Réinitialiser l'index et mettre à jour les colonnes 'row' et 'point_n'
        df_full.reset_index(inplace=True)
        df_full.rename(columns={'index': 'date'}, inplace=True)

        start_row = points[0]['row'] if points else 0
        df_full['row'] = range(start_row, start_row + len(df_full))
        df_full['point_n'] = df_full.index

        # Filtrer pour conserver uniquement les valeurs d'origine
        new_values = df_full[['row', 'value']].dropna(subset=['value']).to_dict('records')
        adjusted_points = [{'date': x['date'].to_pydatetime(), 'row': x['row'], 'point_n': x['point_n']} for x in df_full.drop(columns=['value']).to_dict('records')]

        return adjusted_points, new_values

    def _getValuesByCurve(self, curve):
        result = []
        value = None
        for point in self.points:
            cell = self.ws.cell(row=point["row"], column=curve["column"])
            if cell.value is not None and cell.value != "#REF!":
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        value = self.evaluator.evaluate(self.sheet_name+"!"+cell.coordinate)
                    if "%" in cell.value:
                        value = float(cell.value.replace("%", ""))/100
                elif isinstance(cell.value, float) or isinstance(cell.value, int):
                    value = cell.value
                result.append({
                    "row": point["row"],
                    "value": value
                })
        return result 

    def _evaluateInterpolation(self, curve):
        result = []
        values = self._getValuesByCurve(curve) # Récupère les valeurs de la courbe non nulles

        # Interpolate values
        if (curve["interpolation"] == "CONST" and len(values) > 0) or (curve["interpolation"] is None and len(values) == 1):
            result = [next(item["value"] for item in values if item["value"] is not None)] * self.num_points
        elif len(values) >= 2:
            # Adjust points if necessary
            xs, ys = [v["row"] for v in values], [v["value"] for v in values]           
            interp1d = self.log_interp1d(xs, ys) if curve["interpolation"] == "LOG" else interpolate.interp1d(xs,ys, fill_value="extrapolate")

            if self._determine_current_frequency(self.points) != self.getSimulationFrequency():
                self.points, values = self._adjust_points_frequency(self.points, values, self.getSimulationFrequency())

            result = [float(interp1d(point["row"])) for point in self.points]
        
        if len(result) == 0:
            return []
        if len(result) != self.num_points:
            raise Exception(f"Curve '{curve['curve_name']}' has {len(result)} points instead of {self.num_points} on sheet '{self.sheet_name}'") 
        return result

    # Init Functions

    def _define_num_points(self):
        frequency = self.getSimulationFrequency()
        duration_in_years = int(self.getSummaryByName(self.SIMULATION_END_NAME)["summary_value"].year) - int(self.getSummaryByName(self.SIMULATION_STARTDATE_NAME)["summary_value"].year)
        if frequency in self.FREQ_MULTIPLIER:
            self.duration_in_years = duration_in_years
            self.num_points = duration_in_years * self.FREQ_MULTIPLIER[frequency]
        else:
            raise Exception(f"Frequency {frequency} not recognized")
        
    def loadSheet(self):
        """
            Return JSON storage of the input
        """

        if self.sheet_name.startswith(self.DELIMITER_SHEET_UNFOLLOW):
            logger.info(f"loadSheet('{self.sheet_name}') : SKIPPED")
            return False

        if self.isOperationSheet():
            logger.info(f"loadSheet('{self.sheet_name}') : Formulas_Sheet")
            self.operations = self._generateOperations()
            return True

        if self.isConstantSheet():
            logger.info(f"loadSheet('{self.sheet_name}') : Constants_Sheet")
            self.constants = self._generateConstants()
            return True

        if self.isSummarySheet():
            logger.info(f"loadSheet('{self.sheet_name}') : Summary_Sheet")
            self._generateSummary()
            return True

        logger.info(f"loadSheet('{self.sheet_name}') : Curves_Sheet")

        # Get number of points
        self._define_num_points()
       
        # Get all elements in worksheet
        self.metadatas = self._generateMetaData()
        self.points = self._generate_points_with_dates()
        self.curves = self._generateCurves()
        
        if self.isCurvesSheet() and self.metadatas == {}:
            raise Exception("Metadatas are missing...{}".format(self.sheet_name))
        
        # Interpolate values      
        for curve in self.curves:
            curve["values"] = self._evaluateInterpolation(curve)       

        return True
    
    # Add Functions

    def addCurve(self, specifcation_name, values, unit):
        """
        Add new Curve to the analyzer, if curve exists we do nothing
        """

        if values is None or len(values) != self.num_points:
            raise Exception("Error on values point for curve '{}'".format(specifcation_name))
        
        if self.getCurveByName(specifcation_name) is not None:
            return None

        next_free_column = self.curves[-1]["column"]+1

        self.curves.append({
            "column": next_free_column,
            "curve_name": self._clean_string(specifcation_name),
            "values": values, 
            "unit": unit if unit is not None else "",
            "interpolation": self.determine_interpolation_type(values)
        })
    
    def addSummary(self, summary_name, value, unit):
        if self.getSummaryByName(summary_name) is not None:
            return None

        self.summary.append({
            "summary_name": self._clean_string(summary_name),
            "summary_value": value,
            "unit": unit if unit is not None else ""
        })

    # Access Functions

    def getCurveByName(self, name):
        res = next((item for item in self.curves if self._clean_string(item["curve_name"].lower()) == self._clean_string(name.lower())), None)
        if res is not None :
            res["specificiation_name"] = self._clean_string(res["curve_name"])
        return res
        
    def getConstantByName(self, name):
        result = iter([item for item in list(self.constants.values())[0] if self._clean_string(item["constant_name"].lower()) == self._clean_string(name.lower())])
        return next(result, None)
    
    def getConstantByCategoryAndName(self, category, name):
        if category in self.constants:
            return next(iter([item for item in self.constants[category] if self._clean_string(item["constant_name"].lower()) == self._clean_string(name.lower())]), None)

    def getSummaryByName(self, name):
        result = iter([item for item in self.summary if item["summary_name"].lower() == name.lower()])
        return next(result, None)

    def getSimulationFrequency(self):
        return self.getSummaryByName(self.SIMULATION_FREQUENCY_NAME)["summary_value"].lower() if self.getSummaryByName(self.SIMULATION_FREQUENCY_NAME) is not None else "year"
    def getOperationByName(self, name):
        result = iter([item for item in self.operations if item["operation_name"].lower() == name.lower()])
        return next(result, None)                        

    def getCategory(self):
        if self.CATEGORY in self.metadatas:
            return str(self.metadatas[self.CATEGORY])
        return None
    
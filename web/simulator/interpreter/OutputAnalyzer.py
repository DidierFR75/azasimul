import os
import re
import random
import pathlib
import zipfile
import datetime
import copy
import logging
import pandas as pd
import numpy as np
import scipy as sp
from scipy import interpolate
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from pycel import ExcelCompiler
from openpyxl import Workbook, load_workbook
from anytree import Node, find, findall
from copy import copy, deepcopy
from faker import Faker
from .Helper import Helper

import logging
import os
try:
    logging.config.dictConfig({
    'version': 1,
    'disable_existing_loggers': False,
    'formatters': {
        'console': {
            'format': '%(name)-12s %(levelname)-8s %(message)s'
        },
        'file': {
            'format': '%(asctime)s %(name)-12s %(levelname)-8s %(message)s'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'formatter': 'console'
        },
        'file': {
            'level': 'DEBUG',
            'class': 'logging.FileHandler',
            'formatter': 'file',
            'filename': './logs/log.log'
        }
    },
    'loggers': {
        '': {
            'level': 'DEBUG',
            'handlers': ['console', 'file']
        }
    }
})
    os.makedirs('./logs', exist_ok=True)
    logger = logging.getLogger(__name__)
except Exception as e:
    print(e)
class OutputAnalyzer:
    """
    The `OutputAnalyzer` class is responsible for analyzing and manipulating data in an Excel sheet. 
    It provides methods for converting values based on filters, formatting values based on units, copying cell styles, inserting data according to transformer functions, unmerging cells, and finding and replacing annotated values with curve values.

    Methods:
        - `__init__(self, wb, sheet_name, path, interpreter)`: Initializes an instance of `OutputAnalyzer` with the given workbook, sheet name, file path, and interpreter.
        - `convertFilter(self, value, unit, filter)`: Converts a value based on a filter and unit.
        - `formatByUnit(self, val, unit)`: Formats a value based on its unit.
        - `copyCellStyle(self, cell, new_cell)`: Copies the style of a cell to a new cell.
        - `isInterpretable(self, value)`: Checks if a value is interpretable.
        - `insertTransformer(self, cell, is_lines_needed_already_insert)`: Inserts data according to a transformer function.
        - `unmergeCell(self, cell)`: Unmerges a cell.
        - `findAndReplaceAnnotateValues(self)`: Finds and replaces annotated values with curve values.
        - `save(self, path)`: Saves the modified workbook to the specified path.

    Fields:
        - `EXPRESSION`: Regular expression pattern for matching variable expressions in output cells.
        - `FUNCTION`: Dictionary mapping function names to their corresponding keywords.
        - `FUNCTION_TRANSFORMER`: Dictionary mapping function names to their corresponding transformer functions.
        - `FILTERS_DISPATCH`: Dictionary mapping filter categories to their corresponding dispatch functions.
        - `UNIT_FORMATS`: Dictionary mapping unit names to their corresponding formatting functions.
        - `evaluator`: Instance of `ExcelCompiler` for evaluating Excel formulas.
        - `sheet_name`: Name of the sheet being analyzed.
        - `wb`: Workbook object being analyzed.
        - `ws`: Worksheet object being analyzed.
        - `tree`: Tree object representing the Excel file structure.
        - `interpreter`: Instance of `ExcelInterpreter` for interpreting Excel files.
    """
    VAR_EXPRESSION = r'\[[ \w().|+\-{}]+\]' # expression of a var in output's cell
    
    FUNCTION = {
        "for": "FOR:",
        "hfor": "HFOR:"
    }

    FUNCTION_TRANSFORMER = {
        'for': ["INDEX", "DATEPOINT"]
    }

    FILTERS_DISPATCH = {
        "category": {}
    }

    UNIT_FORMATS = {
        "date" : lambda x: x.strftime('%Y-%m-%d'),
        "$": lambda x: '{:,.2f}'.format(x),
        "€": lambda x: '{:,.2f}'.format(x),
        "cost": lambda x: '{:,.2f}'.format(x),
        "%": lambda x: "{:.2%}".format(x)
    }   

    def __init__(self, wb, sheet_name, path, interpreter) -> None:
        self.evaluator = ExcelCompiler(filename=path)
        self.sheet_name = sheet_name
        self.wb = wb
        self.ws = self.wb[sheet_name]
        self.tree = interpreter.tree
        self.interpreter = interpreter
        self.FILTERS_DISPATCH["category"] = {item: lambda x: x for item in self.tree.root.categories}

    def getNumberOfYears(self):
        start = self.tree.root.analyzer.getSummaryByName(Helper.SIMULATION_STARTDATE_NAME)["summary_value"]
        end = self.tree.root.analyzer.getSummaryByName(Helper.SIMULATION_END_NAME)["summary_value"]
        delta = relativedelta(end, start)

        return delta.years
    
    def getNumberOfPoints(self):
        frequency = self.tree.root.analyzer.getSummaryByName(Helper.SIMULATION_FREQUENCY_NAME)["summary_value"].lower() if self.tree.root.analyzer.getSummaryByName(Helper.SIMULATION_FREQUENCY_NAME) is not None else None
        nb_points = (self.getNumberOfYears()+1) * Helper.FREQ_MULTIPLIER[frequency] if frequency in Helper.FREQ_MULTIPLIER else self.getNumberOfYears()
        return nb_points

    def convertFilter(self, value, unit, filter):
        """
        Convert value by it filter
        Ex : 01/01/2022|year = 2022
        """
        if unit.lower() in self.FILTERS_DISPATCH and filter.lower() in self.FILTERS_DISPATCH[unit]:
            return self.FILTERS_DISPATCH[unit.lower()][filter.lower()](value)
        return value

    def formatByUnit(self, val, unit):
        """
        Format val according to his unit by UNIT_FORMATS function
        """
        if unit is not None and val != "" and unit != "" and unit in list(self.UNIT_FORMATS.keys()):
            try:
                if unit != "date":
                    val = float(val)
                return deepcopy(self.UNIT_FORMATS[unit](val))
            except Exception as e:
                raise Exception("Unit problem : ", e, unit, val)
        return val

    def copyCellStyle(self, cell, new_cell):
        """
        Return new_cell with the style of cell
        """
        if cell is not None and cell.has_style and new_cell is not None:
            try:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            except Exception as e:
                raise Exception(e)

        return new_cell

    def isInterpretable(self, value):
        """
        Check if it's an interpretable value and return true if it is else false
        """
        if value is None or not isinstance(value, str) or value == "" or value == " " or value == "$":
            return False
        
        # Check if value start with one value of self.FUNCTION
        valid_func = next(iter([True for key_func, value_func in self.FUNCTION.items() if value.startswith(value_func)]), False)
        
        try:
           re.search(value, self.VAR_EXPRESSION)
        except:
            raise Exception(value, self.VAR_EXPRESSION)
             
        if re.search(value, self.VAR_EXPRESSION) or valid_func:
            return True
            
        return False

    def _randomize_date(self, date):
        # Générer un nouveau mois et un nouveau jour aléatoires
        random_month = random.randint(1, 12)
        # Générer un jour valide pour le mois
        if random_month == 2:  # Février
            # Tenir compte des années bissextiles pour février
            if (date.year % 4 == 0 and date.year % 100 != 0) or (date.year % 400 == 0):
                random_day = random.randint(1, 29)
            else:
                random_day = random.randint(1, 28)
        elif random_month in [4, 6, 9, 11]:  # Avril, Juin, Septembre, Novembre
            random_day = random.randint(1, 30)
        else:  # Les autres mois
            random_day = random.randint(1, 31)
        
        # Créer une nouvelle date avec l'année originale, le nouveau mois et le nouveau jour
        new_date = date.replace(month=random_month, day=random_day)
        return new_date

    def insertTransformer(self, cell, for_already_insert, values):
        """
        Insert data according to the transformer function and return True if it's done else False
    
        :param cell: The cell in which the transformer function is found.
        :type cell: Cell object
        :param for_already_insert: A flag indicating whether data has already been inserted for the transformer function.
        :type for_already_insert: bool
        :return: True if data is inserted based on the transformer function, False otherwise.
        :rtype: bool
        """
        if isinstance(cell.value, str) and cell.value.startswith(self.FUNCTION["for"]):
            l = [item for item in self.FUNCTION_TRANSFORMER["for"] if cell.value.endswith(item)]
            if l != []:
                l = l[0]
            
                # Add date if YEAR else add index
                unit = "date" if l == "DATEPOINT" else None

                if not for_already_insert:
                    for i in range(1, len(values)):
                        self.ws.insert_rows(cell.row+i)

                if l == "INDEX":
                    values = range(0, len(values))

                self.ws.cell(row=cell.row, column=cell.column).value = self.formatByUnit(values[0], unit)
                for i in range(1, len(values)):
                    self.ws.cell(row=cell.row+i, column=cell.column).value = self.formatByUnit(values[i], unit)
                    self.copyCellStyle(cell, self.ws.cell(row=cell.row+i, column=cell.column))
                return True
        elif isinstance(cell.value, str) and cell.value.startswith(self.FUNCTION["hfor"]):
            l = [item for item in self.FUNCTION_TRANSFORMER["for"] if cell.value.endswith(item)]
            if l != []:
                l = l[0]
            
                # Add date if YEAR else add index
                unit = "date" if l == "DATEPOINT" else None

                if l == "INDEX":
                    values = range(0, len(values))

                self.ws.cell(row=cell.row, column=cell.column).value = self.formatByUnit(values[0], unit)
                for i in range(1, len(values)):
                    self.ws.cell(row=cell.row, column=cell.column+i).value = self.formatByUnit(values[i], unit)
                    self.copyCellStyle(cell, self.ws.cell(row=cell.row, column=cell.column+i))
                return True
        return False

    def unmergeCell(self, cell):
        # Check if cell is MergedCell and unmerge it
        for mergecells in self.ws.merged_cells.ranges:
            pass
    
    def findNode(self, treeRoot, category, scope, verbose=False):
        """
        Find the first node in a tree structure based on its category and scope.

        Args:
            treeRoot (Node): The root node of the tree structure.
            category (str): The category of the node to find.
            scope (str): The scope of the node to find.
            verbose (bool, optional): Whether to log errors if multiple nodes are found. Defaults to False.

        Returns:
            Node: The first matching node found in the tree based on the given category and scope.

        Raises:
            Exception: If the node with the given category and scope is not found in the tree.
        """
        catCode = category.lower()
        nodes = None
        if scope:
            nodes = findall(treeRoot, lambda node: node.name.lower() == catCode and hasattr(node, 'category') and node.category == scope)
            if not nodes and scope == "root":
                nodes = findall(treeRoot, lambda node: node.name.lower() == catCode and not hasattr(node, 'category'))
            if len(nodes)>1 and verbose:
                logger.error(f"!!! nodes_according_to_category({catCode}) Scope({scope}).length()>1")
        if not nodes:
            nodes = findall(treeRoot, lambda node: node.name.lower() == catCode) 
            if not nodes:
                raise Exception(f"Cannot find sheet '{catCode}' in the tree... with Scope '{scope}'")

        node = nodes[0] if nodes else nodes
        return node

    def findAndReplaceAnnotateValues(self):
        """
        Find and replace all annotate's values by their curve's value
        """
        for merge_cell in self.ws.merged_cells.ranges:
            self.ws.unmerge_cells(str(merge_cell))

        freq_rule = Helper.PD_FREQ_MULTIPLIER[self.tree.root.analyzer.getSimulationFrequency()]
        timestamps = pd.date_range(
            start=self.tree.root.analyzer.getSummaryByName(Helper.SIMULATION_STARTDATE_NAME)["summary_value"],
            end=self.tree.root.analyzer.getSummaryByName(Helper.SIMULATION_END_NAME)["summary_value"]+relativedelta(years=1),
            freq=freq_rule
        ).to_list() 
        date_range = [timestamp.to_pydatetime() for timestamp in timestamps]

        for row in self.ws:
            for_already_insert = False
            for cell in row:
                # add new row if not already did
                if self.isInterpretable(cell.value):
                    
                    if self.insertTransformer(cell, for_already_insert, date_range):
                        for_already_insert = True
                        continue

                    matches = re.finditer(self.VAR_EXPRESSION, cell.value)
                    for match in matches:
                        node = None
                        full_match = match.group(0)
                        inner_content = full_match.strip('[]')  # Remove the brackets

                        variable_part, filter = self.splitVariablePart(inner_content)

                        line_number = None
                        if '{' in variable_part:
                            # Extract the variable name and line number if present
                            variable_part, line_info = re.match(r'(.*?){(\d+)}', variable_part).groups()
                            line_number = int(line_info)

                        attr = variable_part.split(".")

                        if len(attr) > 1:                            
                            node = self.findNode(self.tree.root, attr[0], filter, verbose=False)

                            if node is not None:
                                data = None
                                val = {}

                                if node.analyzer.isCurvesSheet():
                                    data = node.analyzer.getCurveByName(attr[1])                
                                    if data is not None:
                                        # Interprete FOR directive according to val
                                        if cell.value.startswith(self.FUNCTION["for"]) and [item for item in self.FUNCTION_TRANSFORMER["for"] if cell.value.endswith(item)] == []:

                                            # if for not previously added and insert necessary row
                                            if not for_already_insert:
                                                for_already_insert = True
                                                for i in range(1, self.getNumberOfPoints()):
                                                    self.ws.insert_rows(cell.row+i)
                                            
                                            # Add values to each cell
                                            for i in range(0, self.getNumberOfPoints()):
                                                self.ws.cell(row=cell.row+i, column=cell.column).value = self.formatByUnit(data["values"][i], data["unit"])
                                                self.copyCellStyle(cell, self.ws.cell(row=cell.row+i, column=cell.column))
                                            continue
                                        
                                        # Interprete HFOR directive according to val
                                        if cell.value.startswith(self.FUNCTION["hfor"]) and [item for item in self.FUNCTION_TRANSFORMER["for"] if cell.value.endswith(item)] == []:
                                            
                                            # Add values to each cell
                                            for i in range(0, self.getNumberOfPoints()):
                                                self.ws.cell(row=cell.row, column=cell.column+i).value = self.formatByUnit(data["values"][i], data["unit"])
                                                self.copyCellStyle(cell, self.ws.cell(row=cell.row, column=cell.column+i))
                                            continue

                                        # Get data value
                                        try:
                                            if data["interpolation"] == "CONST":
                                                val = data["values"][0]
                                            else:
                                                val = data["values"][line_number] if line_number is not None and line_number <= len(data["values"]) else data["values"][0]
                                        except:
                                            raise Exception("Can't access data values ", data)
                                            
                                if node.analyzer.isConstantSheet() and len(attr) > 2:
                                    data = node.analyzer.getConstantByCategoryAndName(attr[1], attr[2])
                                    val = data["value"] if data is not None else None                    

                                if node.analyzer.isSummarySheet():
                                    data = node.analyzer.getSummaryByName(attr[1])
                                    val = data["summary_value"] if data is not None else None

                                val = self.formatByUnit(val, data["unit"]) if data is not None and "unit" in data else val
                                
                                cell.value = val if val != {} and val is not None else ""

        for merge_cell in self.ws.merged_cells.ranges:
            self.ws.merge_cells(str(merge_cell))

    def splitVariablePart(self, inner_content):
        parts = inner_content.split('|')
        variable_part = parts[0]
        filter = parts[1].lower() if len(parts) > 1 else None
        return variable_part, filter

    def getNodeAndLineNumber(self, variable_part, filter):
        line_number = None
        if '{' in variable_part:
            variable_part, line_info = re.match(r'(.*?){(\d+)}', variable_part).groups()
            line_number = int(line_info)

        attr = variable_part.split(".")
        node = self.findNode(self.tree.root, attr[0], filter, verbose=False) if len(attr) > 1 else None
        return node, line_number

    def getVariableValue(self, node, variable_part, line_number):
        attr = variable_part.split(".")
        if node.analyzer.isCurvesSheet():
            data = node.analyzer.getCurveByName(attr[1])
            if data and data["interpolation"] != "CONST":
                try:
                    return data["values"][line_number] if line_number else data["values"][0]
                except IndexError:
                    raise Exception("Invalid line number for curve data")
            return data["values"][0] if data else None

        if node.analyzer.isConstantSheet() and len(attr) > 2:
            data = node.analyzer.getConstantByCategoryAndName(attr[1], attr[2])
            return data["value"] if data else None

        if node.analyzer.isSummarySheet():
            data = node.analyzer.getSummaryByName(attr[1])
            return data["summary_value"] if data else None

        return None

    def save(self, path):
        self.wb.save(path)
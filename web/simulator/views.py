from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.core.mail import send_mail, BadHeaderError
from django.http import HttpResponse
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.models import User
from django.template.loader import render_to_string
from django.db.models.query_utils import Q
from django.utils.http import urlsafe_base64_encode
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.http import Http404

from .forms import NewUserForm, SimulationForm
from .models import Simulation, SimulationInput, MODEL_INPUT_PATH, MODEL_OUTPUT_PATH
from .interpreter import rejectXlsFile, FileChecker, SheetOutputGenerator, SheetInterpreter, folder_zip, reject_file

from openpyxl import Workbook, load_workbook

import os
import zipfile
import shutil
import datetime


@login_required(login_url="simulator:login")
def index(request):
    simulations = Simulation.objects.all()
    return render(request, "dashboard/index.html", {"simulations": simulations})

def WS_update_fromForm(simul_id, form):
    # Modify excel file according to the database properties
    input_path = Simulation.getPath(simul_id,'inputs')
    for model_file in os.listdir(input_path):
        if reject_file(model_file):
            continue
        wb_path = input_path+"/"+model_file
        wb = load_workbook(wb_path)
        wb.iso_dates = True
        changesNbr = 0
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            for composition in ws["A"]:
                if composition.value is not None:
                    val = form.cleaned_data.get(composition.value.lower().replace(" ", "_"))
                    if composition.value.lower().replace(" ", "_") in form.cleaned_data.keys() and ws.cell(row=composition.row, column=composition.column+1).value != val:
                        if isinstance(val, datetime.datetime):
                            val = datetime.datetime.strftime(val, "%Y-%m-%d %H:%M:%S")
                        ws.cell(row=composition.row, column=composition.column+1).value = val
                        changesNbr += 1
        if changesNbr:
            wb.save(wb_path)

# Modify excel file according to the database properties
            # input_files = settings.MEDIA_ROOT + "/inputs/simulation_" + str(simulation.id)
            # for model_file in os.listdir(input_files):
            #     wb = load_workbook(input_files+"/"+model_file)
            #     wb.iso_dates = True
            #     if "Summary" in wb.sheetnames:
            #         ws = wb["Summary"]
            #         for composition in ws["A"]:
            #             if composition.value is not None:
            #                 val = form.cleaned_data.get(composition.value.lower().replace(" ", "_"))
            #                 if composition.value.lower().replace(" ", "_") in form.cleaned_data.keys() and ws.cell(row=composition.row, column=composition.column+1).value != val:
            #                     val = datetime.datetime.strftime(val, "%Y-%m-%d %H:%M:%S") if isinstance(val, datetime.datetime) else val

            #                     ws.cell(row=composition.row, column=composition.column+1).value = val
            #                     wb.save(input_files+"/"+model_file)

@login_required(login_url="simulator:login")
def new(request):
    if request.method == "POST":
        form = SimulationForm(request.POST, request.FILES)
        if form.is_valid():

            # Add current User in request
            simulation = form.save(commit=False)
            simulation.user = request.user
            
            # Files Handler
            files = request.FILES.getlist('input_files')
            simul_inputs = []
            for f in files:
                # Temporary save file
                path = settings.MEDIA_ROOT+"/tmp/"+str(f)
                default_storage.save(path, ContentFile(f.read()))
                
                # Check if file has specification format or if it has summary sheet
                fc = FileChecker(path)
                fc.checkForSpecFormat()

                # Add additionnal data in simulation
                if fc.summary:
                    for summary in fc.summary:
                        summary_name = summary["summary_name"].lower().replace(" ", "_")
                        if hasattr(simulation, summary_name) and (getattr(simulation, summary_name) is None or getattr(simulation, summary_name) != summary["summary_value"]):
                            # If value in form different of value in files, we modify the file's values according to the form's values
                            if form.cleaned_data.get(summary_name) != summary["summary_value"] and form.cleaned_data.get(summary_name) != None and form.cleaned_data.get(summary_name) != "" and form.cleaned_data.get(summary_name) != " ":                                
                                summary_name = form.cleaned_data.get(summary_name)
                                                        
                            setattr(simulation, summary_name, summary["summary_value"])
                        
                if fc.non_accepted != []:
                    messages.error(request, "The following sheets were not take into account: "+ ','.join(fc.non_accepted))

                # Create file object
                simul_inputs.append(SimulationInput(input_file=f))

                os.remove(path)

            # Create simulation object and inputs objects
            simulation.save()
            simulation.createPaths()
            path = Simulation.getPath(simulation.id, 'inputs')+'/'
            os.makedirs(path, exist_ok=True)

            for simul_input in simul_inputs:
                simul_input.simulation = simulation
                simul_input.save()

            WS_update_fromForm(simulation.id, form)
            messages.success(request, "The simulation has been registered.")
            return redirect("simulator:index")
        
        message = ""
        if form.errors:
            for field in form:
                for error in field.errors:
                    message = message + error + ', '
        messages.error(request, "Error: " + message)

    form = SimulationForm()
    return render(request, 'dashboard/new.html', {"simulation_form": form})

@login_required(login_url="simulator:login")
def edit(request, id):
    simulation = get_object_or_404(Simulation, id=id)
    if request.method == "POST":
        form = SimulationForm(request.POST, request.FILES, instance=simulation)
        if form.is_valid():
            form.save()
            
            # Files Handler
            files = request.FILES.getlist('input_files')
            # Delete previous inputs
            if len(files) > 0:
                simulation.simulation_input.all().delete()
                # Add new inputs
                for f in files:
                    path = settings.MEDIA_ROOT+"/tmp/"+str(f)
                    default_storage.save(path, ContentFile(f.read()))
                    
                    # Check if file has specification format or if it has summary sheet
                    fc = FileChecker(path)
                    fc.checkForSpecFormat()

                    # Add additionnal data in simulation
                    if fc.summary is not None:
                        for summary in fc.summary:
                            summary["summary_name"] = summary["summary_name"].lower().replace(" ", "_")
                            if hasattr(simulation, summary["summary_name"]) and (getattr(simulation, summary["summary_name"]) is None or getattr(simulation, summary["summary_name"]) != summary["summary_value"]):
                                # If value in form different of value in files, we modify the file's values according to the form's values
                                if form.cleaned_data.get(summary["summary_name"]) != summary["summary_value"] and form.cleaned_data.get(summary["summary_name"]) != None and form.cleaned_data.get(summary["summary_name"]) != "" and form.cleaned_data.get(summary["summary_name"]) != " ":
                                    summary["summary_value"] = form.cleaned_data.get(summary["summary_name"])
                                setattr(simulation, summary["summary_name"], summary["summary_value"])

                    SimulationInput.objects.create(input_file=f, simulation=simulation)
                            
                    os.remove(path)

            WS_update_fromForm(simulation.id, form)
            messages.success(request, "The simulation has been updated.")
            return redirect("simulator:index")
        
        message = ""
        if form.errors:
            for field in form:
                for error in field.errors:
                    message = message + error + ', '
        messages.error(request, "An error appear : " + message)
    
    form = SimulationForm(instance=simulation)
    return render(request, "dashboard/edit.html", {
        "edit_form": form,
        "simulation": simulation
        })
    
@login_required(login_url="simulation:login")
def delete(request, id):
    simulation = get_object_or_404(Simulation, id=id)
    simulation.delete()
    messages.success(request, f'The simulation #{id} "{simulation.project_name}" has been deleted.')
    return redirect('simulator:index')

def response_zip_file(zip_path, zip_fn, removeAfterDownload=False):
    """Zip all output file and serve to download"""
    from django.http import FileResponse
    zip_file = open(zip_path, 'rb')
    response = FileResponse(zip_file, as_attachment=True, filename=zip_fn)

    # response = HttpResponse(
    #     zip_file,
    #     headers={
    #         "Content-Type": "application/zip",
    #         "Content-Disposition": f'attachment; filename="{zip_fn}"',
    #     },
    # )
    # response = HttpResponse(zip_file, content_type='application/zip')
    # response['Content-Disposition'] = f'attachment; filename="{zip_fn}"'
    if removeAfterDownload:
        os.remove(zip_path)
    return response

def response_zip(include_inputs=False):
    pass

@login_required(login_url='simulation:login')
def doCompute(request, simul_id, downloadInputs=False):
    simulation = get_object_or_404(Simulation, id=simul_id) 
    input_path = Simulation.getPath(simul_id,'inputs')

    if downloadInputs:
        zip_fn = f"SimAZA_{simulation.project_name}_inputs_{simulation.created_at.strftime('%Y-%m-%d')}"
        zip_path = folder_zip(input_path, zip_fn)
    else:
        # Copy default .xlsx files (Common Operations, Financial...)
        for fn in os.listdir(MODEL_INPUT_PATH):
            if rejectXlsFile(fn):
                continue
            src = MODEL_INPUT_PATH + fn
            dst = f"{input_path}/{fn}"
            shutil.copyfile(src, dst)

        interpreter = SheetInterpreter(input_path)
        interpreter.evaluate()
        # Generate output files
        generator = SheetOutputGenerator(interpreter, MODEL_OUTPUT_PATH)
        generator.analyzeAllOutputSheet()
        result_path = Simulation.getPath(simul_id,'outputs')
        zip_fn = f"SimAZA_{simulation.project_name}_{simulation.created_at.strftime('%Y-%m-%d')}"
        zip_path = generator.generate(result_path, zip_fn)

    zip_fn += ".zip"
    return response_zip_file(zip_path, zip_fn)

@login_required(login_url='simulation:login')
def downloadData(request, simul_id):
    return doCompute(request, simul_id, downloadInputs=True)

@login_required(login_url='simulation:login')
def downloadOneData(request, id, namefile):
    simulation = get_object_or_404(Simulation, id=id) 

    # path = settings.MEDIA_ROOT+"/inputs/simulation_"+str(simulation.id)
    path = Simulation.getPath(simulation.id, 'inputs')+'/'

    # Copy /operations in media/input/simulation_id to take into account default operations
    for model_file in os.listdir(path):
        if model_file == namefile:
            # Zip all output file and serves to download
            zip_file = open(path+"/"+namefile, 'rb')
            response = HttpResponse(zip_file, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="'+namefile+'"'

            return response
    raise Exception("File not exist in this simulation")

@login_required(login_url='simulation:login')
def listDownloadData(request, id):
    path = Simulation.getPath(id, 'inputs')+'/'
    try: files = os.listdir(path)
    except: files=[]
    if not files:
        return HttpResponse("No files")
    return render(request, 'dashboard/listdatas.html', {"input_files": files, "simulation_id": id})

# Add Questions/Constants page
@login_required(login_url="simulator:login")
def index_co(request, type):
    model = MODEL_INPUT_PATH if type == 'input' else MODEL_OUTPUT_PATH

    return render(request, 'co/index.html', {
        "models": [fn for fn in os.listdir(model) if not rejectXlsFile(fn)],
        "type": type
    })

@login_required(login_url="simulator:login")
def new_co(request, type):
    model = MODEL_INPUT_PATH if type == 'input' else MODEL_OUTPUT_PATH
    if request.method == "POST":
        # Save file uploaded
        files = request.FILES.getlist('files')
        for f in files:
            default_storage.save(model+str(f), ContentFile(f.read()))
        
        messages.success(request, "The new operations/constants has been register !")
        return redirect("simulator:index_co", type=type)
        
    return render(request, 'co/new.html')

@login_required(login_url="simulator:login")
def download_co(request, type, name):
    model = MODEL_INPUT_PATH if type == 'input' else MODEL_OUTPUT_PATH

    path = model+name
    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(path)
            return response
    raise Http404
   
@login_required(login_url="simulator:login")
def delete_co(request, type, name):
    model = MODEL_INPUT_PATH if type == 'input' else MODEL_OUTPUT_PATH

    path = model+name
    if os.path.exists(path):
        os.remove(path)
        return redirect('simulator:index_co', type=type)
    raise Http404


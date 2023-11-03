from unicodedata import category
# from openpyxl.writer.excel import save_virtual_workbook
import os
import re
import random
import pathlib
import zipfile
import math
import datetime
import copy
import pandas as pd
import numpy as np
import scipy as sp
from scipy import interpolate
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from pycel import ExcelCompiler
from openpyxl import Workbook, load_workbook
from statistics import mean
from anytree import Node, RenderTree, find, Resolver, PostOrderIter, findall
from copy import copy, deepcopy
from faker import Faker

import logging
try:
    logging.config.dictConfig({
    'version': 1,
    'disable_existing_loggers': False,
    'formatters': {
        'console': {
            'format': '%(name)-12s %(levelname)-8s %(message)s'
        },
        'file': {
            'format': '%(asctime)s %(name)-12s %(levelname)-8s %(message)s'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'formatter': 'console'
        },
        'file': {
            'level': 'DEBUG',
            'class': 'logging.FileHandler',
            'formatter': 'file',
            'filename': './logs/log.log'
        }
    },
    'loggers': {
        '': {
            'level': 'DEBUG',
            'handlers': ['console', 'file']
        }
    }
})
    os.makedirs('./logs', exist_ok=True)
    logger = logging.getLogger(__name__)
except Exception as e:
    print(e)

def rejectXlsFile(fn):
    if fn.startswith(".") or fn.startswith(InputAnalyzer.DELIMITER_SHEET_UNFOLLOW) or not fn.endswith('.xlsx'):
        return True
    return False

def reject_file(file_path):
    fn = os.path.basename(file_path)
    if rejectXlsFile(fn):
        return True
    return False

def folder_zip(folderPath, zip_fn):
    """
    Create <zip_fn>.zip of a folder and return path
    """
    directory = pathlib.Path(folderPath)
    destination = f"{directory.parent.absolute()}/{zip_fn}.zip"

    with zipfile.ZipFile(destination, mode="w") as archive:
        for file_path in directory.iterdir():
            if reject_file(file_path):
                continue
            archive.write(file_path, arcname=file_path.name)
    return destination

"""
The `InputAnalyzer` class is responsible for analyzing and processing data from an Excel worksheet. 
It can identify different types of sheets (operation, constant, summary, and curves) and extract relevant information from each type. 
It can also perform interpolation on curves and store the results.

Main functionalities:
- Identify and load different types of sheets (operation, constant, summary, and curves)
- Extract metadata, points, curves, constants, summary, and operations from the sheet
- Perform interpolation on curves
- Add new curves, constants, summary, and operations to the analyzer

Methods:
- loadSheet(): Loads and analyzes the sheet, extracting relevant information based on the sheet type
- addCurve(specification_name, value, unit, interpolation): Adds a new curve to the analyzer
- addConstant(name, value, unit): Adds a new constant to the analyzer
- addSummary(name, value, unit): Adds a new summary to the analyzer
- addOperation(name, operation, unit): Adds a new operation to the analyzer
- getCurveByName(name): Retrieves a curve by its name
- getConstantByName(name): Retrieves a constant by its name
- getSummaryByName(name): Retrieves a summary by its name
- getOperationByName(name): Retrieves an operation by its name
- getCategory(): Retrieves the category metadata of the sheet

Fields:
- BASE_ELEMENTS_ROW: The row number where the base elements are located
- UNIT_ROW: The row number where the units are located
- CURVE: The row number where the curve interpolations are located
- METADATA_COL: The column letter where the metadata is located
- POINT_N_COL: The column letter where the point numbers are located
- DATE_COL: The column letter where the dates are located
- PRODUCT_NAME: The metadata name for the product type
- PRODUCT_PARENT: The metadata name for the product subtype
- CATEGORY: The metadata name for the category
- DELIMITER_SHEET_UNFOLLOW: The delimiter used to skip sheets
- CONSTANT_SHEETNAME: The name of the constant sheet
- SUMMARY_SHEETNAME: The name of the summary sheet
- OPERATION_SHEETNAME: The name of the operation sheet
- evaluator: An instance of the ExcelCompiler class for evaluating Excel formulas
- sheet_name: The name of the sheet being analyzed
- ws: The worksheet object being analyzed
- curves: A list of dictionaries representing the curves in the sheet
- points: A list of dictionaries representing the points in the sheet
- metadatas: A dictionary representing the metadata in the sheet
- operations: A dictionary representing the operations in the sheet
- constants: A dictionary representing the constants in the sheet
- summary: A list of dictionaries representing the summary in the sheet
- fake: An instance of the Faker class for generating fake data
"""
class InputAnalyzer:
    BASE_ELEMENTS_ROW = 17 # Location of base's elements
    UNIT_ROW = 18 # Location of units
    CURVE = 19 # Location of Curve's interpolations
    METADATA_COL = "A"
    POINT_N_COL = "B"
    DATE_COL = "D"
    PRODUCT_NAME = "Product-Type"
    PRODUCT_PARENT = "SubType"
    CATEGORY = "Category"
    DELIMITER_SHEET_UNFOLLOW = "_"

    CONSTANT_SHEETNAME = "Constants"
    SUMMARY_SHEETNAME = "Summary" 
    OPERATION_SHEETNAME = "operation"

    def __init__(self, ws, sheet_name, path) -> None:
        self.evaluator = ExcelCompiler(filename=path)
        self.sheet_name = sheet_name
        self.ws = ws

        self.curves = []
        self.points = []
        self.metadatas = {}
        self.operations = {}
        self.constants = {}
        self.summary = []
        self.fake = Faker()
    
    ### Util's functions

    def log_interp1d(self, xx, yy, kind='linear', small_value=1e-10):
        """
        Return the log interpolation on 1 dimension
        """
        yy_adjusted = np.maximum(yy, small_value)
        logx = np.log10(xx)
        logy = np.log10(yy_adjusted)
        lin_interp = interpolate.interp1d(logx, logy, kind=kind, fill_value="extrapolate")
        log_interp = lambda zz: np.power(10.0, lin_interp(np.log10(zz)))
        return log_interp

    def evaluate(self, cell):
        """
        Return the evaluation's value of a given cell (rounded by 3)
        """
        eval = self.evaluator.evaluate(self.sheet_name+"!"+cell.coordinate)
        if isinstance(eval, float):
            eval = round(eval, 3)
        return eval
    
    def clean_string(self, text):
        return str(text).replace("\n", "").replace("\t", "").lstrip()

    ### 4 Sheet types

    def isOperationSheet(self):
        if self.OPERATION_SHEETNAME.lower() in self.sheet_name.lower():
            return True
        return False

    def isConstantSheet(self):
        if self.CONSTANT_SHEETNAME.lower() in self.sheet_name.lower():
            return True
        return False
    
    def isSummarySheet(self):
        if self.SUMMARY_SHEETNAME.lower() in self.sheet_name.lower():
            return True
        return False

    def isCurvesSheet(self):
        if self.curves != [] and self.points != []:
            return True
        return False

    # Generator's functions

    def _generatePointsWithDates(self):
        """
            Return dict (row_number, point_no, date) of all points and associate's date
        """
        
        points = [
            {"row": point.row, "point_n": self.evaluate(point), "date": self.ws[self.DATE_COL+str(point.row)].value } 
            for point in self.ws[self.POINT_N_COL] 
            if point.value is not None and (isinstance(point.value, (int, float)) or (isinstance(point.value, str) and point.value.startswith("=")))
        ]
        
        valid_points = [ pt for pt in points if pt["date"] is not None ]
        # Check if exist at least 2 dates in points
        if len(valid_points) >= 2:
            x = list(range(len(valid_points)))
            try:
                y = list(map(datetime.datetime.timestamp, [pt["date"] for pt in valid_points]))
            except Exception as e:
                raise e
            interp1d = interpolate.interp1d(x, y, fill_value="extrapolate") 
            interpolated_timestamps = interp1d(x)
            for pt, interpolated_timestamp in zip(valid_points, interpolated_timestamps):
                pt["date"] = datetime.datetime.fromtimestamp(interpolated_timestamp)
            for pt, interpolated_timestamp in zip(points, interpolated_timestamps):
                if pt["date"] is None:
                    pt["date"] = datetime.datetime.fromtimestamp(interpolated_timestamp)
            return points
        else:
            return self._populateByFrequency(points)

    def _generateCurves(self):
        """
        Return a sorted (by ASC) dict (column, curve, unit, interpolation) at BASE_ELEMENTS_ROW for a given sheetname
        Ignore the firsts 2 elements because they always not belong to Curves
        """
        return sorted([{
            "column": be.column, 
            "curve_name": self.clean_string(be.value),
            "values": None,
            "unit": self.ws.cell(row=self.UNIT_ROW, column=be.column).value, 
            "interpolation": self.ws.cell(row=self.CURVE, column=be.column).value
        } for be in self.ws[self.BASE_ELEMENTS_ROW] if be.value is not None][2:], key=lambda x: x["column"])

    def _generateMetaData(self):
        """
        Return array of tuples (row_id, metadata_name, metadata_value) for a given sheetname
        """
        return {self.clean_string(cell.value): self.ws['B'][cell.row-1].value 
            for cell in self.ws[self.METADATA_COL] 
            if (cell.value is not None and self.ws['B'][cell.row-1].value is not None and cell.row < self.BASE_ELEMENTS_ROW)}

    def _generateConstants(self):
        """
        Return all global constant in the sheet
        """
        compositions = [(cmp.row, cmp.value) for cmp in self.ws["A"] if cmp.value is not None]

        result = {}
        for index, cmp in enumerate(compositions):
            tmp = []            
            if(index != len(compositions)-1):
                last_row = compositions[index+1][0]-1
            else:
                last_row = len(self.ws["B"])

            for x in range(cmp[0]+1, last_row):
                if self.ws["B"+str(x)].value is not None:
                    tmp.append({
                        "constant_name": self.clean_string(self.ws["B"+str(x)].value),
                        "value": self.evaluate(self.ws["C"+str(x)]),
                        "unit": self.evaluate(self.ws["D"+str(x)])
                    })
            
            result[cmp[1]] = tmp

        return result

    def _generateSummary(self):
        for composition in self.ws["A"]:
            if composition.value is not None and self.ws.cell(row=composition.row, column=composition.column+1).value is not None:
                self.addSummary(composition.value, self.ws.cell(row=composition.row, column=composition.column+1).value, self.ws.cell(row=composition.row, column=composition.column+2).value)

    def _generateOperations(self):
        """
        Return operations
        """
        items = [(it.row, it.value) for it in self.ws["A"] if it.value is not None]

        result = {}
        for index, fcn in enumerate(items):
            tmp = []            
            if(index != len(items)-1):
                last_row = items[index+1][0]-1
            else:
                last_row = len(self.ws["B"])

            for x in range(fcn[0]+1, last_row):
                if self.ws["B"+str(x)].value is not None:
                    tmp.append({
                        "operation_name": self.clean_string(self.ws["B"+str(x)].value),
                        "operation": self.evaluate(self.ws["C"+str(x)]),
                        "unit": self.clean_string(self.ws["D"+str(x)].value) 
                    })
            
            result[fcn[1]] = tmp

        return result

    def _populateByFrequency(self, points, granularity="year", X=1):
        """
        Populates the given list of points with interpolated dates based on a reference date.

        Parameters:
        - points (list): A list of dictionaries where each dictionary has a "date" key.
                        The list should have at least one point with a non-null date.
        - granularity (str): The unit for interpolation. Supported values are "year", "month", 
                        "day", "hour", and "minute".
        - X (int): The multiplier for the granularity.

        The function will find the first non-null date in the points list and use it as a reference.
        It will then populate missing dates in the list based on the given granularity and multiplier.

        Returns:
        - None: The function modifies the input list in-place.
        """

        # Mapping granularities to their respective relativedelta functions and random generators using faker
        granularity_map = {
            "year": lambda base_date, diff: base_date + relativedelta(years=random.randint(1, diff)),
            "month": lambda base_date, diff: base_date + relativedelta(months=random.randint(1, diff)),
            "day": lambda base_date, diff: base_date + timedelta(days=random.randint(1, diff)),
            "hour": lambda base_date, diff: base_date + timedelta(hours=random.randint(1, diff)),
            "minute": lambda base_date, diff: base_date + timedelta(minutes=random.randint(1, diff))
        }

        # Create the reference date as the first one found
        ref_date = next(({"index": index, "date": point["date"]} for index, point in enumerate(points) if point["date"]), None)

        if ref_date:
            for index, point in enumerate(points):
                if point["date"] is None:
                    point["date"] = granularity_map[granularity](ref_date, X)
        
        return points

    def _getValuesByCurve(self, curve):
        result = []
        value = None
        for point in self.points:
            cell = self.ws.cell(row=point["row"], column=curve["column"])
            if cell.value is not None and cell.value != "#REF!":
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        value = self.evaluator.evaluate(self.sheet_name+"!"+cell.coordinate)
                    if "%" in cell.value:
                        value = float(cell.value.replace("%", ""))/100
                elif isinstance(cell.value, float) or isinstance(cell.value, int):
                    value = cell.value
                result.append({
                    "row": point["row"],
                    "value": value
                })
        return result 

    def _evaluateInterpolation(self, curve):
        result = []
        values = self._getValuesByCurve(curve)
        
        if (curve["interpolation"] == "CONST" and len(values) > 0) or (curve["interpolation"] is None and len(values) == 1):
            result = [next(item["value"] for item in values if item["value"] is not None)] * len(self.points)
        elif len(values) >= 2:
            xs, ys = [v["row"] for v in values], [v["value"] for v in values]           
            if curve["interpolation"] == "LOG":
                interp1d = self.log_interp1d(xs, ys)
            else:
                interp1d = interpolate.interp1d(xs,ys, fill_value="extrapolate")

            # Add result of interpolation for each point
            result = [float(interp1d(point["row"])) for point in self.points]
        return result

    # Init Functions

    def loadSheet(self):
        """
            Return JSON storage of the input
        """

        if self.sheet_name.startswith(self.DELIMITER_SHEET_UNFOLLOW):
            logger.info(f"loadSheet('{self.sheet_name}') : SKIPPED")
            return False

        if self.isOperationSheet():
            logger.info(f"loadSheet('{self.sheet_name}') : Formulas_Sheet")
            self.operations = self._generateOperations()
            return True

        if self.isConstantSheet():
            logger.info(f"loadSheet('{self.sheet_name}') : Constants_Sheet")
            self.constants = self._generateConstants()
            return True

        if self.isSummarySheet():
            logger.info(f"loadSheet('{self.sheet_name}') : Summary_Sheet")
            self._generateSummary()
            return True

        logger.info(f"loadSheet('{self.sheet_name}') : Curves_Sheet")

        # Get all elements in worksheet
        self.metadatas = self._generateMetaData()
        self.points = self._generatePointsWithDates()
        self.curves = self._generateCurves()
        
        if self.isCurvesSheet() and self.metadatas == {}:
            raise Exception("Metadatas are missing...{}".format(self.sheet_name))

        # Interpolate values  
        for curve in self.curves:
            curve["values"] = self._evaluateInterpolation(curve)       
        
        return True
    
    # Add Functions

    def addCurve(self, specifcation_name, value, unit, interpolation="CONST"):
        """
        Add new Curve to the analyzer, if curve exists we do nothing
        """

        if self.getCurveByName(specifcation_name) is not None:
            return None

        next_free_column = self.curves[-1]["column"]+1

        self.curves.append({
            "column": next_free_column,
            "curve_name": self.clean_string(specifcation_name),
            "values": [value] * len(self.points), 
            "unit": unit if unit is not None else "",
            "interpolation": interpolation
        })
    
    def addSummary(self, summary_name, value, unit):
        if self.getSummaryByName(summary_name) is not None:
            return None

        self.summary.append({
            "summary_name": self.clean_string(summary_name),
            "summary_value": value,
            "unit": unit if unit is not None else ""
        })

    # Access Functions

    def getCurveByName(self, name):
        res = next((item for item in self.curves if self.clean_string(item["curve_name"].lower()) == self.clean_string(name.lower())), None)
        if res is not None :
            res["specificiation_name"] = self.clean_string(res["curve_name"])
        return res
        
    def getConstantByName(self, name):
        result = iter([item for item in list(self.constants.values())[0] if self.clean_string(item["constant_name"].lower()) == self.clean_string(name.lower())])
        return next(result, None)
    
    def getConstantByCategoryAndName(self, category, name):
        if category in self.constants:
            return next(iter([item for item in self.constants[category] if self.clean_string(item["constant_name"].lower()) == self.clean_string(name.lower())]), None)

    def getSummaryByName(self, name):
        result = iter([item for item in self.summary if item["summary_name"].lower() == name.lower()])
        return next(result, None)

    def getOperationByName(self, name):
        result = iter([item for item in self.operations if item["operation_name"].lower() == name.lower()])
        return next(result, None)                        

    def getCategory(self):
        if self.CATEGORY in self.metadatas:
            return str(self.metadatas[self.CATEGORY])
        return None

"""
The `SheetTree` class is responsible for creating a tree structure to organize and analyze sheets in Excel workbooks.

Attributes:
    path (str): The path to the folder containing the Excel workbooks.
    root (Node): The root node of the formula tree.
    all_sheet (dict): A dictionary with the file names as keys and a list of sheet analyzers as values.
    operation_sheets (list): A list of sheet analyzers for the operation sheets.
"""
class SheetTree:
    def __init__(self, path) -> None:
        """
        Initializes a new instance of the SheetTree class.
        
        Args:
            path (str): The path to the folder containing the Excel workbooks.
        """
        self.path = path
        
        self.root = Node("root")
        self.root.name = "Summary"
        self.root.categories = {}

        self.all_sheet = None
        self.operation_sheets = []

    def readAllSheetsFromFolder(self, folder):
        """
        Reads all the sheets from a folder and returns a dictionary with the file names as keys and a list of sheet analyzers as values.
        
        Args:
            folder (str): The path to the folder containing the Excel workbooks.
        
        Returns:
            dict: A dictionary with the file names as keys and a list of sheet analyzers as values.
        """
        result = {}

        # Load all workbooks
        all_files = next(os.walk(folder), (None, None, []))[2]
        all_files = [ fn for fn in all_files if not rejectXlsFile(fn) ]
        all_wks = { file: load_workbook(folder +'/'+ file) for file in all_files }
        
        # Create dict with file: {sheetname: analyzer}
        for file, wb in all_wks.items():
            result[file] = []
            for sheet_name in wb.sheetnames:
                if sheet_name.startswith(InputAnalyzer.DELIMITER_SHEET_UNFOLLOW):
                    logger.info(f"Sheet('{sheet_name}') : SKIPPED")
                    continue
                analyzer = InputAnalyzer(wb[sheet_name], sheet_name, folder + '/' + file)
                if analyzer.loadSheet():
                    result[file].append((sheet_name, analyzer,))

        return result

    def mapSheetsToFormulaTree(self, path=None):
        """
        Maps the sheets to the formula tree structure.
        
        Args:
            path (str, optional): The path to the folder containing the Excel workbooks. If not provided, uses the default path.
        """
        if not path:
            path = self.path
        nodes = []
        self.all_sheet = self.readAllSheetsFromFolder(path)
        # Create all nodes
        for _file, wbSheets in self.all_sheet.items():
            for sheet_name, analyzer in wbSheets:
                if analyzer.isOperationSheet():
                    self.operation_sheets.append(analyzer)
                    continue
            
                if analyzer.isConstantSheet():
                    Node(sheet_name, analyzer=analyzer, parent=self.root)
                    continue

                if analyzer.isSummarySheet():
                    if not hasattr(self.root, "analyzer"):
                        self.root.analyzer = analyzer
                    else:
                        # Delete summary value with same key that root_summary
                        for root_summary in self.root.analyzer.summary:
                            for summary in analyzer.summary:
                                if root_summary["summary_name"].lower() == summary["summary_name"].lower():
                                    del summary
                        # Merge summaries values
                        for summary in analyzer.summary:
                            self.root.analyzer.summary.append(summary)
                                                        
                    continue

                if analyzer.metadatas == {}:
                    continue
                
                # last Case is a "Curves" sheet
                # Get parent name if exists
                parentName = analyzer.metadatas[analyzer.PRODUCT_PARENT] if (analyzer.PRODUCT_PARENT in analyzer.metadatas) else None
                    
                productType = analyzer.metadatas[analyzer.PRODUCT_NAME]
                node = Node(productType, analyzer=analyzer)
                
                # Get and add category to self.root.categories if exist
                category = analyzer.getCategory()
                if category is not None:
                    self.root.categories[category.lower()] = category.upper()+":"
                    node.category = category.lower()
                                            
                nodes.append( (parentName, productType, node ) )
        
        # Add parent for all nodes
        for element in nodes:
            if element[0] is None: # if no parentName
                element[2].parent = self.root
            else:
                i = [i for i, v in enumerate(nodes) if v[1] == element[0] and element[2].category == v[2].category]
                if i != []:
                    element[2].parent = nodes[i[0]][2]

def findNode(treeRoot, category, scope, verbose=False):
    catCode = category.lower()
    nodes = None
    if scope:
        nodes = findall(treeRoot, lambda node: node.name.lower() == catCode and hasattr(node, 'category') and node.category == scope)
        if not nodes and scope == "root":
            nodes = findall(treeRoot, lambda node: node.name.lower() == catCode and not hasattr(node, 'category'))
        if len(nodes)>1 and verbose:
            logger.error(f"!!! nodes_according_to_category({catCode}) Scope({scope}).length()>1")
    if not nodes:
        nodes = findall(treeRoot, lambda node: node.name.lower() == catCode) 
        if not nodes:
            raise Exception(f"Cannot find sheet '{catCode}' in the tree... with Scope '{scope}'")

    node = nodes[0] if nodes else nodes
    return node

"""
The `SheetInterpreter` class is responsible for interpreting and evaluating formulas in Excel workbooks. 
It replaces variables and functions in the formulas with their corresponding values and then evaluates the formulas to obtain the final results.

Main functionalities:
- Replaces variables and functions in formulas with their corresponding values
- Evaluates the formulas to obtain the final results
- Updates the summary and curves sheets in the Excel workbooks with the evaluated results

Methods:
- __init__(self, folder): Initializes a new instance of the SheetInterpreter class with the specified folder path.
- findOperation(self, category, operation_name): Finds an operation by its category and operation name.
- convertFilter(self, value, unit, filter): Converts a value based on a filter.
- replaceOneVarByValue(self, word, default_node, category, scope): Replaces a variable in a formula with its corresponding value.
- replaceAllVarsByValue(self, opStr, default_node, category, scope): Replaces all variables in a formula with their corresponding values.
- replaceFcnByVar(self, opStr, category, scope): Replaces functions in a formula with their corresponding values.
- mapOperationValues(self, list_operations, category, scope): Maps the values in a list of operations by replacing variables and functions with their corresponding values.
- operationParser(self): Parses the operations by replacing functions with variables and mapping the values.
- evaluate(self): Evaluates the formulas by replacing variables and functions with their values and updating the summary and curves sheets.

Fields:
- tree: The SheetTree object that represents the tree structure of the Excel workbooks.
- node_categories: A list of the categories of the nodes in the tree.
- operations: A dictionary that stores the evaluated operations categorized by node category.
"""
class SheetInterpreter:

    FILTERS_DISPATCH = {
        "date" : {
            "year" : lambda x: x.year,
            "month": lambda x: x.month,
            "day": lambda x: x.day
        }
    }

    FCN_EXPR = '\{[ \_\(\)\-\|\.a-zA-Z0-9]+\}'
    VAR_EXPR = '\[[ \_\(\)\|\-\+a-zA-Z0-9\.]+\]'

    def __init__(self, folder) -> None:
        self.tree = SheetTree(folder)
        self.tree.mapSheetsToFormulaTree()
        self.node_categories = list(self.tree.root.categories.keys()) # list(map(lambda x: x.lower(), list(self.tree.root.categories.keys())))
        self.operations = {cat: [] for cat in self.node_categories}
        self.operations["root"] = []
    
    # Utils functions

    def findOperation(self, category, operation_name):
        """
        Find an operation by it category and it operation_name 
        """
        for analyzer in self.tree.operation_sheets:
            for analyzer_category, operations in analyzer.operations.items():
                if analyzer_category.lower() == category.lower():
                    return next((operation for operation in operations if operation["operation_name"].lower() == operation_name.lower()), None)
        return None

    def convertFilter(self, value, unit, filter):
        """
        Convert value by it filter
        Ex : datetime(01/01/2022)|year = 2022
        """
        try:
            func = self.FILTERS_DISPATCH[unit.lower()][filter.lower()]
            value = func(value)
        except: pass
        return value

    # Functions to search and replace variables and functions
    
    def replaceOneVarByValue(self, word, default_node, category, scope):
        """
        Replace Word by his Variable Value in the according worksheet
        """

        if word is None or category is None:
            raise Exception("replaceOneVarByValue need all paramaters fill...")
        
        correct_word = word.replace("[", "").replace("]", "")
        attr = correct_word.split('.')

        if len(attr) == 1:
            attr.insert(0, category)

        if len(attr) > 1:
            node = findNode(self.tree.root, attr[0], scope)
            # nodes = findall(self.tree.root, lambda node: node.name.lower() == attr[0].lower()) 

            # if not nodes:
            #     raise Exception(f"Cannot find sheet '{attr[0]}' in the tree... with category '{category}'")

            # if len(nodes) > 1:
            #     node = find(self.tree.root, lambda n: hasattr(n, "category") and n.category == dst and n.name.lower() == attr[0].lower())
            #     if not node:
            #         node = nodes[0]
            # else:
            #     node = nodes[0]
        else:
            node = default_node

        # It is a constant
        if len(attr) == 3 and node.analyzer.isConstantSheet():
            constant = node.analyzer.getConstantByCategoryAndName(attr[1], attr[2])
            if constant is not None:
                return constant["value"]
            
            return None

        if node.analyzer.isSummarySheet():
            cw = attr[1].split('|')

            summary = node.analyzer.getSummaryByName(cw[0])

            if summary:
                if len(cw) == 2:
                    return self.convertFilter(summary["summary_value"], summary["unit"], cw[1])
                    
                return summary["summary_value"]
            return None

        if node.analyzer.isCurvesSheet():
            spec = node.analyzer.getCurveByName(attr[1])
            
            if spec is not None:
                if spec["interpolation"] == "CONST":
                    return spec["values"][0]
                else:
                    # make an average of each interpolate's values 
                    return mean(spec["values"])
            else:
                raise Exception('Error: replaceOneVarByValue()', word, node, category, scope, attr, node.analyzer.curves)
        
        return None
    
    def replaceAllVarsByValue(self, opStr, default_node, category, scope):
        # Replace first all vars [] by value in result
        for m in re.finditer(self.VAR_EXPR, opStr):
            opStr = opStr.replace(m.group(0), str(self.replaceOneVarByValue(m.group(0), default_node, category, scope)))
        return opStr

    def replaceFcnByVar(self, opStr, category, scope):
        """
            Replace all fcn {} by value
        """
        matches = re.finditer(self.FCN_EXPR, opStr)
        while matches is not None:
            for match in matches:
                according_op = None
                wks = None

                fcn_name = match.group(0).replace("{", "").replace("}", "").strip()
                attr = fcn_name.split('.')

                # if operation exist in list operation, add the value of it in it
                if len(attr) > 2:
                    raise Exception("unknown function syntax for:", attr)
                
                if len(attr) == 1:
                    attr.insert(0, category)
                
                according_op = self.findOperation(attr[0], attr[1])

                if according_op is not None:
                    related_nodes = findall(self.tree.root, lambda n: n.name.lower() == attr[0].lower())
                else:
                    raise Exception("prb :", attr, match.group(0), fcn_name)
                
                if according_op is not None and related_nodes != ():
                    for wks in related_nodes:                            
                        # Transform all {} in children by interpretable {}
                        accStr = according_op["operation"]
                        for m in re.finditer(self.FCN_EXPR, accStr):
                            rpl = m.group(0).replace("{", "").replace("}", "").strip()

                            if len(rpl.split(".")) == 1:
                                accStr = accStr.replace(m.group(0), "{"+attr[0]+"."+rpl+"}")

                        accStr = self.replaceAllVarsByValue(accStr, wks, attr[0], scope)
                        
                        opStr = opStr.replace(match.group(0), "("+ accStr+ ")")
                        try:
                            opStr = str(round(eval(opStr), 10))
                        except:
                            pass
                else:
                    raise Exception("A problem is in :", attr, match.group(0))    

            if re.search(self.FCN_EXPR, opStr) is not None:
                matches = re.finditer(self.FCN_EXPR, opStr)
            else:
                matches = None
        try:
            opStr = str(round(eval(opStr), 10))
        except:
            pass
        return opStr

    def mapOperationValues(self, list_operations, category, scope):
        try:
            default_node = findNode(self.tree.root, category, scope, verbose=False)

            if default_node:                  
                copy_operations = copy.deepcopy(list_operations)    
                
                for index, op in enumerate(copy_operations):                            
                    if op["operation"] is None:
                        raise Exception("Operation can't be null :", op)
            
                    opStr = self.replaceAllVarsByValue(op["operation"], default_node, category, scope)
                    op["operation"] = self.replaceFcnByVar(opStr, category, scope)
                    
                    op["node_category"] = default_node

                    copy_operations[index] = op
                    
                return copy_operations
        except Exception as e:
            logger.error(e)
            print(e)
            raise e

        
        return None
        
    def operationParser(self):
        """
        Replace all Formula {} by Var [] while it's present in string of all operations 
        """
        for analyzer in self.tree.operation_sheets:
            for category, l_operations in analyzer.operations.items():

                if category is None or l_operations is None or not isinstance(category, str):
                    raise Exception('ReplaceFcnByVar needs category and operations')

                # Map values in self.operations
                for dst, stored_operations in self.operations.items():
                    values = self.mapOperationValues(l_operations, category, dst)

                    if values:
                        stored_operations.append(values) 

    # Render functions
    def evaluate(self):
        """
        Replace all [] Expression by their Values to be evaluate next
        """
        
        # Search all {} operations and replace by []
        self.operationParser()

        # Eval all operations
        for dst, list_operations in self.operations.items():
            for operations in list_operations:
                    for operation in operations:
                        try:
                            # Attention si l'user met rm -rf * par exemple !!
                            # logging.log(operation["operation"])
                            operation["operation"] = round(eval(str(operation["operation"])), 4)
                            # logging.log(operation["operation"])
                            if operation["node_category"].analyzer.isSummarySheet():
                                operation["node_category"].analyzer.addSummary(operation["operation_name"], operation["operation"], operation["unit"])
                            elif operation["node_category"].analyzer.isCurvesSheet():
                                operation["node_category"].analyzer.addCurve(operation["operation_name"], operation["operation"], operation["unit"], "CONST")
                        except Exception as e:
                            raise Exception("Error for evaluation of operations", operation, e)

"""
The `OutputAnalyzer` class is responsible for analyzing and manipulating data in an Excel sheet. 
It provides methods for converting values based on filters, formatting values based on units, copying cell styles, inserting data according to transformer functions, unmerging cells, and finding and replacing annotated values with curve values.

Methods:
    - `__init__(self, wb, sheet_name, path, interpreter)`: Initializes an instance of `OutputAnalyzer` with the given workbook, sheet name, file path, and interpreter.
    - `convertFilter(self, value, unit, filter)`: Converts a value based on a filter and unit.
    - `formatByUnit(self, val, unit)`: Formats a value based on its unit.
    - `copyCellStyle(self, cell, new_cell)`: Copies the style of a cell to a new cell.
    - `isInterpretable(self, value)`: Checks if a value is interpretable.
    - `insertTransformer(self, cell, for_already_insert)`: Inserts data according to a transformer function.
    - `unmergeCell(self, cell)`: Unmerges a cell.
    - `findAndReplaceAnnotateValues(self)`: Finds and replaces annotated values with curve values.
    - `save(self, path)`: Saves the modified workbook to the specified path.

Fields:
    - `EXPRESSION`: Regular expression pattern for matching variable expressions in output cells.
    - `FUNCTION`: Dictionary mapping function names to their corresponding keywords.
    - `FUNCTION_TRANSFORMER`: Dictionary mapping function names to their corresponding transformer functions.
    - `FILTERS_DISPATCH`: Dictionary mapping filter categories to their corresponding dispatch functions.
    - `UNIT_FORMATS`: Dictionary mapping unit names to their corresponding formatting functions.
    - `evaluator`: Instance of `ExcelCompiler` for evaluating Excel formulas.
    - `sheet_name`: Name of the sheet being analyzed.
    - `wb`: Workbook object being analyzed.
    - `ws`: Worksheet object being analyzed.
    - `tree`: Tree object representing the Excel file structure.
    - `interpreter`: Instance of `ExcelInterpreter` for interpreting Excel files.
"""
class OutputAnalyzer:

    EXPRESSION = '\[[ \_\(\)\|\-\+a-zA-Z0-9\.]+\]' # expression of a var in output's cell
    
    FUNCTION = {
        "for": "FOR:",
    }

    FUNCTION_TRANSFORMER = {
        'for': ["INDEX", "DATEPOINT"]
    }

    FILTERS_DISPATCH = {
        "category": {}
    }

    UNIT_FORMATS = {
        "date" : lambda x: x.strftime('%Y-%m-%d'),
        "$": lambda x: '{:,.2f}'.format(x),
        "€": lambda x: '{:,.2f}'.format(x),
        "cost": lambda x: '{:,.2f}'.format(x),
        "%": lambda x: "{:.2%}".format(x)
    }   

    def __init__(self, wb, sheet_name, path, interpreter) -> None:
        self.evaluator = ExcelCompiler(filename=path)
        self.sheet_name = sheet_name
        self.wb = wb
        self.ws = self.wb[sheet_name]
        self.tree = interpreter.tree
        self.interpreter = interpreter
        self.FILTERS_DISPATCH["category"] = {item: lambda x: x for item in self.tree.root.categories}

    def convertFilter(self, value, unit, filter):
        """
        Convert value by it filter
        Ex : 01/01/2022|year = 2022
        """
        if unit.lower() in self.FILTERS_DISPATCH and filter.lower() in self.FILTERS_DISPATCH[unit]:
            return self.FILTERS_DISPATCH[unit.lower()][filter.lower()](value)
        return value

    def formatByUnit(self, val, unit):
        """
        Format val according to his unit by UNIT_FORMATS function
        """
        if unit is not None and val != "" and unit != "" and unit in list(self.UNIT_FORMATS.keys()):
            try:
                return copy.deepcopy(self.UNIT_FORMATS[unit](val))
            except Exception as e:
                raise Exception("Unit problem : ", e, unit, val)
        return val

    def copyCellStyle(self, cell, new_cell):
        """
        Return new_cell with the style of cell
        """
        if cell is not None and cell.has_style and new_cell is not None:
            try:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = copy.copy(cell.number_format)
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)
            except Exception as e:
                raise Exception(e)

        return new_cell

    def isInterpretable(self, value):
        """
        Check if it's an interpretable value and return true if it is else false
        """
        if value is None or not isinstance(value, str) or value == "" or value == " " or value == "$":
            return False
        
        # Check if value start with one value of self.FUNCTION
        valid_func = next(iter([True for key_func, value_func in self.FUNCTION.items() if value.startswith(value_func)]), False)
        
        try:
           re.search(value, self.EXPRESSION)
        except:
            raise Exception(value, self.EXPRESSION)
             
        if re.search(value, self.EXPRESSION) or valid_func:
            return True
            
        return False

    def _randomize_date(self, date):
        # Générer un nouveau mois et un nouveau jour aléatoires
        random_month = random.randint(1, 12)
        # Générer un jour valide pour le mois
        if random_month == 2:  # Février
            # Tenir compte des années bissextiles pour février
            if (date.year % 4 == 0 and date.year % 100 != 0) or (date.year % 400 == 0):
                random_day = random.randint(1, 29)
            else:
                random_day = random.randint(1, 28)
        elif random_month in [4, 6, 9, 11]:  # Avril, Juin, Septembre, Novembre
            random_day = random.randint(1, 30)
        else:  # Les autres mois
            random_day = random.randint(1, 31)
        
        # Créer une nouvelle date avec l'année originale, le nouveau mois et le nouveau jour
        new_date = date.replace(month=random_month, day=random_day)
        return new_date

    def insertTransformer(self, cell, for_already_insert):
        """
        Insert data according to the transformer function and return True if it's done else False
        """

        if isinstance(cell.value, str) and cell.value.startswith(self.FUNCTION["for"]):
            l = [item for item in self.FUNCTION_TRANSFORMER["for"] if cell.value.endswith(item)]
            if l != []:
                l = l[0]
                
                start = self.tree.root.analyzer.getSummaryByName("Start")["summary_value"]
                end = self.tree.root.analyzer.getSummaryByName("End")["summary_value"]
                delta = relativedelta(end, start)
                
                # Add date if DATEPOINT else add index
                values = list(map(lambda x: self._randomize_date(start + relativedelta(years=x)) if l == "DATEPOINT" else x+1, [item for item in range(0, delta.years+1)]))
                unit = "date" if l == "DATEPOINT" else None

                if not for_already_insert:
                    for i in range(1, len(values)):
                        self.ws.insert_rows(cell.row+i)
                
                self.ws.cell(row=cell.row, column=cell.column).value = self.formatByUnit(values[0], unit)
                for i in range(1, len(values)):
                    self.ws.cell(row=cell.row+i, column=cell.column).value = self.formatByUnit(values[i], unit)
                    self.copyCellStyle(cell, self.ws.cell(row=cell.row+i, column=cell.column))
                return True
        return False

    def unmergeCell(self, cell):
        # Check if cell is MergedCell and unmerge it
        for mergecells in self.ws.merged_cells.ranges:
            pass

    def findAndReplaceAnnotateValues(self):
        """
        Find and replace all annotate's values by their curve's value
        """
        
        for row in self.ws:
            for_already_insert = False
            for cell in row:
                # add new row if not already did
                if self.isInterpretable(cell.value):
                    
                    if self.insertTransformer(cell, for_already_insert):
                        for_already_insert = True
                        continue

                    matches = re.finditer(self.EXPRESSION, cell.value)
                    for match in matches:
                        node = None
                        m = match.group(0).replace("[", "").replace("]", "").strip()
                        attr = m.split(".")

                        if len(attr) > 1:
                            # Check if filter exist
                            filter = attr[1].split("|")
                            attr[1] = filter[0]

                            node = findNode(self.tree.root, attr[0], filter[1].lower() if len(filter) > 1 else None, verbose=False)
                         
                            if node is not None:
                                data = None
                                val = {}

                                if node.analyzer.isCurvesSheet():
                                    data = node.analyzer.getCurveByName(attr[1])
                                                        
                                    if data is not None:
                                        # Interprete FOR directive according to val
                                        if cell.value.startswith(self.FUNCTION["for"]) and [item for item in self.FUNCTION_TRANSFORMER["for"] if cell.value.endswith(item)] == []:
                                            start = self.tree.root.analyzer.getSummaryByName("Start")["summary_value"]
                                            end = self.tree.root.analyzer.getSummaryByName("End")["summary_value"]
                                            delta = relativedelta(end, start)
                                            nb_points = delta.years

                                            # if for not previously added and insert necessary row
                                            if not for_already_insert:
                                                for_already_insert = True
                                                for i in range(1, nb_points):
                                                    self.ws.insert_rows(cell.row+i)
                                            
                                            # Add values to each cell
                                            for i in range(0, nb_points):
                                                self.ws.cell(row=cell.row+i, column=cell.column).value = self.formatByUnit(data["values"][i], data["unit"])
                                                self.copyCellStyle(cell, self.ws.cell(row=cell.row+i, column=cell.column))
                                            continue
                                        
                                        # Get data value
                                        if data["interpolation"] == "CONST":
                                            try:
                                                val = data["values"][0]
                                            except:
                                                raise Exception("No value found for : ", data)
                                        else:
                                            try:
                                                val = mean(data["values"])
                                            except:
                                                raise Exception("Can't do a mean of value for ", data)

                                if node.analyzer.isConstantSheet() and len(attr) > 2:
                                    data = node.analyzer.getConstantByCategoryAndName(attr[1], attr[2])
                                    val = data["value"] if data is not None else None                    

                                if node.analyzer.isSummarySheet():
                                    data = node.analyzer.getSummaryByName(attr[1])
                                    val = data["summary_value"] if data is not None else None

                                val = self.formatByUnit(val, data["unit"]) if data is not None and "unit" in data else val
                                
                                cell.value = val if val != {} and val is not None else ""

    def save(self, path):
        self.wb.save(path)

    """
    The `SheetOutputGenerator` class is responsible for generating the final output Excel file by analyzing and replacing variables in the output sheets based on the values obtained from the interpreter.

    Args:
        interpreter (SheetInterpreter): The interpreter object used to evaluate the input file.
        output_path (str): The path to the output files.
    """

"""
The SheetOutputGenerator class is responsible for generating the final output Excel file by replacing variables in the output sheets with their corresponding values obtained from the interpreter.

Args:
    interpreter (Interpreter): The interpreter object used to obtain the values for the variables in the output sheets.
    output_path (str): The path to the output files.

Attributes:
    output_path (str): The path to the output files.
    interpreter (Interpreter): The interpreter object used to obtain the values for the variables in the output sheets.
    all_sheets (dict): A dictionary containing the analyzed output sheets for each output file.
"""
class SheetOutputGenerator:
    
    def __init__(self, interpreter, output_path):
        """
        Initializes the SheetOutputGenerator object with the specified interpreter and output path.

        Args:
            interpreter (Interpreter): The interpreter object used to obtain the values for the variables in the output sheets.
            output_path (str): The path to the output files.
        """
        self.output_path = output_path
        self.interpreter = interpreter
        self.all_sheets = None

    def analyzeAllOutputSheet(self):
        """
        Analyzes all output sheets by creating an OutputAnalyzer object for each sheet in each output file.
        Sets self.all_sheets to a dictionary containing the analyzed output sheets for each output file.
        """
        all_files = next(os.walk(self.output_path), (None, None, []))[2]
        all_wks = {file: load_workbook(self.output_path + file) for file in all_files if file.endswith('.xlsx')}
        
        self.all_sheets = {}
        for file, wb in all_wks.items():
            sheetsDic = {
                sheet_name: OutputAnalyzer(wb, sheet_name, self.output_path + file, self.interpreter)
                for sheet_name in all_wks[file].sheetnames
            }
            self.all_sheets[file] = sheetsDic

    def generate(self, folder, zip_fn):
        """
        Generates the final output Excel file by replacing variables in the output sheets with their corresponding values obtained from the interpreter.

        Args:
            folder (str): The folder to save the generated output files.
            zip_fn (str): The filename of the generated zip file.

        Returns:
            str: The path to the generated zip file.
        """
        from pathlib import Path
        os.makedirs(folder, exist_ok=True)

        count = 0
        for fPath, sheets in self.all_sheets.items():
            count +=1
            fn = Path(fPath).stem
            for sheet_name, analyzer in sheets.items():
                analyzer.findAndReplaceAnnotateValues()
                analyzer.save(f"{folder}/{fn}.xlsx")

        return folder_zip(folder, zip_fn)

"""
This class is responsible for checking the format of a file. 
It loads the file using the `load_workbook` function from the `openpyxl` library and iterates over each sheet in the workbook. 
For each sheet, it creates an instance of the `InputAnalyzer` class to analyze the sheet. 
If the sheet is a summary sheet, it stores the summary data. 
If the sheet is a constant or operation sheet, it adds the sheet name to the list of non-accepted files. 
Finally, it saves the modified workbook.

Fields:
- path: The path of the file to be checked.
- summary: The summary data extracted from the summary sheet.
- non_accepted: A list of sheet names that are not accepted (constant or operation sheets).
- wb: The `Workbook` object representing the loaded workbook.
"""
class FileChecker:

    def __init__(self, path) -> None:
        """
        Initializes the `FileChecker` object with the path of the file to be checked.

        Args:
        - path: The path of the file to be checked.
        """
        self.path = path
        self.summary = None
        self.non_accepted = []
        self.wb = None

    def checkForSpecFormat(self):
        """
        Checks the format of the file. Loads the file using `load_workbook` function, iterates over each sheet, analyzes each sheet using `InputAnalyzer` class, stores the summary data if the sheet is a summary sheet, adds the sheet name to the list of non-accepted files if the sheet is a constant or operation sheet, and saves the modified workbook.
        """
        self.wb = load_workbook(self.path)
        for sheet_name in self.wb.sheetnames:
            analyzer = InputAnalyzer(self.wb[sheet_name], sheet_name, self.path)
            if analyzer.loadSheet():
                if analyzer.isSummarySheet():
                    self.summary = analyzer.summary
                    self.wb.remove(self.wb[sheet_name])
                
                if analyzer.isConstantSheet() or analyzer.isOperationSheet():
                    self.non_accepted.append(sheet_name)
                    self.wb.remove(self.wb[sheet_name])         
        self.wb.save(self.path)